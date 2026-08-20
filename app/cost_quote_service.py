"""原価見積り管理の業務ロジック（Access Flask 版を PostgreSQL 向けに移植）。"""
from __future__ import annotations

from io import BytesIO
import os
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

from app.database import get_connection

def _rec_str(rec, key):
    """詳細クエリ1行 dict から列値を文字列で取得（None は空文字、0 も表示）"""
    if not rec:
        return ""
    v = rec.get(key)
    if v is None:
        return ""
    return str(v)


def _initial_cost_display_str(key, v):
    """初期費用テーブル表示用。数量・単価・金額は long（整数・小数なし）。"""
    if v is None:
        return ""
    if key in ("数量", "単価", "金額"):
        s = str(v).strip().replace(",", "")
        if s == "":
            return ""
        try:
            return str(int(Decimal(s)))
        except (InvalidOperation, ValueError, TypeError, OverflowError):
            return str(v).strip()
    return str(v)


def _initial_cost_rows_for_api(cur, estimate_id):
    """t_原価見積初期費用 の一覧を API／画面更新用 dict のリストで返す"""
    cur.execute(
        "SELECT * FROM t_原価見積初期費用 WHERE 原価見積りID = ?;",
        [estimate_id],
    )
    ic_col_names = [c[0] for c in (cur.description or [])]
    return [
        {k: _initial_cost_display_str(k, v) for k, v in zip(ic_col_names, ic_row)}
        for ic_row in cur.fetchall()
    ]


def _initial_cost_parse_long_for_save(raw, label_jp):
    """登録・更新用: 数量・単価・金額を long として解釈"""
    if raw is None:
        raise ValueError(f"{label_jp}が入力されていません")
    s = str(raw).replace(",", "").strip()
    if s == "":
        raise ValueError(f"{label_jp}が入力されていません")
    try:
        return int(Decimal(s))
    except (InvalidOperation, ValueError, TypeError, OverflowError) as e:
        raise ValueError(f"{label_jp}の値が不正です") from e


def format_cutting_recovery_rate_display(raw):
    """DB値（小数 0.9 等）を画面表示用の数値パーセント（90）へ。空は 90。"""
    if raw is None:
        return "90"
    s = str(raw).strip()
    if s == "":
        return "90"
    try:
        f = float(s.replace(",", ""))
        if 0 <= f <= 1:
            return f"{f * 100:g}"
        return f"{f:g}"
    except ValueError:
        return "90"


def normalize_cutting_recovery_rate_for_db(s):
    """フォーム入力（90%, 90, 0.9）を DB 保存用に正規化（小数の文字列）。"""
    if s is None:
        return ""
    t = str(s).strip().replace("%", "").strip()
    if t == "":
        return ""
    try:
        f = float(t.replace(",", ""))
        if f > 1:
            return str(f / 100.0)
        return str(f)
    except ValueError:
        return ""


def format_yield_rate_display(raw):
    """歩留り: DB の小数（0.1＝10%）を画面の % 表記へ。空は空のまま。"""
    if raw is None:
        return ""
    s = str(raw).strip()
    if s == "":
        return ""
    try:
        f = float(s.replace(",", ""))
        if 0 <= f <= 1:
            return f"{f * 100:g}"
        return f"{f:g}"
    except ValueError:
        return ""


def normalize_yield_rate_for_db(s):
    """歩留り: 画面の %（10）を DB 用の小数（0.1）へ。管理費率と同様。"""
    if s is None:
        return ""
    t = str(s).strip().replace("%", "").strip()
    if t == "":
        return ""
    try:
        f = float(t.replace(",", ""))
        if f > 1:
            return str(f / 100.0)
        return str(f)
    except ValueError:
        return ""


def _format_rate_sum_3(v):
    """賃率計の表示/連携用に小数点以下3桁へ丸めた文字列を返す。"""
    if v is None:
        return ""
    try:
        d = Decimal(str(v)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        return format(d, "f")
    except (InvalidOperation, ValueError, TypeError):
        return ""


EXCEL_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),
    "exceltemplates",
    "原価見積書原本.xlsx",
)

def get_est_calc_page(payload=None):
    """原価見積りID を元に詳細情報を取得して est_calc.html に表示"""
    est_1 = est_2 = est_3 = est_4 = ""
    est_5 = est_6 = est_7 = est_8 = ""
    est_9 = est_10 = est_11 = est_12 = ""
    mat_1 = mat_2 = mat_3 = mat_4 = mat_5 = mat_6 = mat_7 = mat_8 = mat_9 = ""
    mat_10 = mat_11 = mat_12 = mat_13 = mat_14 = mat_15 = mat_16 = mat_17 = mat_18 = mat_19 = ""
    shinchuu_has_row = False
    shinchuu_r1 = None
    shinchuu_r2 = None
    br_1 = br_2 = br_3 = br_4 = br_5 = br_6 = br_7 = br_8 = br_9 = br_10 = ""
    proc_1 = proc_2 = proc_3 = proc_4 = proc_5 = proc_6 = proc_7 = proc_8 = ""
    proc_9 = proc_10 = proc_11 = proc_12 = proc_13 = proc_14 = proc_15 = proc_16 = proc_17 = ""
    proc_18 = proc_19 = proc_20 = proc_21 = proc_22 = proc_23 = proc_24 = proc_25 = proc_26 = proc_27 = ""
    proc_28 = proc_29 = proc_30 = proc_31 = proc_32 = proc_33 = proc_34 = proc_35 = proc_36 = proc_37 = proc_38 = proc_39 = ""
    proc_cb_1 = proc_cb_2 = proc_cb_3 = proc_cb_4 = proc_cb_5 = proc_cb_6 = False
    kensa_1 = kensa_2 = kensa_3 = kensa_4 = kensa_5 = kensa_6 = kensa_7 = kensa_8 = kensa_9 = ""
    kensa_10 = kensa_11 = kensa_12 = kensa_13 = kensa_14 = kensa_15 = kensa_16 = kensa_17 = kensa_18 = kensa_19 = ""
    kensa_20 = kensa_21 = kensa_22 = kensa_23 = kensa_24 = kensa_25 = kensa_26 = kensa_27 = kensa_28 = kensa_29 = kensa_30 = kensa_31 = ""
    kensa_cb_1 = kensa_cb_2 = kensa_cb_3 = kensa_cb_4 = kensa_cb_5 = kensa_cb_6 = False
    est_soryo_box = "1"
    soryo_1 = soryo_2 = soryo_3 = soryo_4 = soryo_5 = soryo_6 = soryo_7 = soryo_8 = soryo_9 = ""
    soryo_10 = soryo_11 = soryo_12 = soryo_13 = soryo_14 = soryo_15 = soryo_16 = soryo_17 = soryo_18 = soryo_19 = ""
    soryo_20 = soryo_21 = soryo_22 = soryo_23 = soryo_24 = soryo_25 = soryo_26 = soryo_27 = soryo_28 = soryo_29 = ""
    soryo_30 = soryo_31 = soryo_32 = soryo_33 = soryo_34 = soryo_35 = ""
    soryo_36 = soryo_37 = soryo_38 = soryo_39 = ""
    f_in_1 = f_in_2 = f_in_3 = ""
    f_out_1 = f_out_2 = f_out_3 = f_out_4 = f_out_5 = f_out_6 = ""
    f_in_b1 = f_in_b2 = f_in_b3 = f_in_b4 = ""
    f_ch_1 = f_ch_2 = f_ch_3 = f_ch_4 = f_ch_5 = f_ch_6 = f_ch_7 = f_ch_8 = f_ch_9 = f_ch_10 = ""
    lot_rows = []
    initial_cost_rows = []
    payload = payload or {}
    estimate_id = str(payload.get("estimate_id") or "").strip()
    zairyo_2_options = []
    kakou_2_options = []
    kakou_5_options = []
    kakou_14_options = []
    kakou_24_options = []
    kakou_32_options = []
    kakou_34_options = []
    kakou_36_options = []
    soryo_2_options = []
    soryo_3_options = []
    soryo_8_options = []
    soryo_13_options = []
    soryo_29_options = []
    rate_default_map = {
        "blast": "",
        "pressfit": "",
        "pre_inspection": "",
        "kensa_auto": "",
        "kensa_numeric": "",
        "kensa_visual": "",
        "kensa_microscope": "",
        "kensa_microgauge": "",
        "kensa_other": "",
    }

    # プルダウン用マスタ取得
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT * FROM t_比重マスタ ORDER BY ID;")
        rows = cur.fetchall()
        zairyo_2_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "specgravity": (r[2] if len(r) > 2 and r[2] is not None else ""),
            }
            for r in rows
        ]

        cur.execute(
            "SELECT * FROM t_材質マスタ;"
        )
        rows = cur.fetchall()
        kakou_2_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "specgravity": (r[2] if len(r) > 2 and r[2] is not None else ""),
                "weight": (r[3] if len(r) > 3 and r[3] is not None else ""),
                "cycle": (r[4] if len(r) > 4 and r[4] is not None else ""),
            }
            for r in rows
        ]

        cur.execute(
            "SELECT ID, 設備名等, [労務費賃率] + [油等賃率] + [電気賃率] + [設備費賃率] + [建屋] + [土地] AS 賃率計 "
            "FROM t_賃率マスタ "
            "WHERE 工程分類 = '未選択' OR 工程分類 = '切削' "
            "ORDER BY ID;"
        )
        rows = cur.fetchall()
        kakou_5_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "rate_sum": _format_rate_sum_3(r[2]),
            }
            for r in rows
        ]

        cur.execute(
            "SELECT ID, 設備名等, [労務費賃率] + [油等賃率] + [電気賃率] + [設備費賃率] + [建屋] + [土地] AS 賃率計 "
            "FROM t_賃率マスタ "
            "WHERE 工程分類 = '未選択' OR 工程分類 = 'バレル' "
            "ORDER BY ID;"
        )
        rows = cur.fetchall()
        kakou_14_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "rate_sum": _format_rate_sum_3(r[2]),
            }
            for r in rows
        ]

        cur.execute(
            "SELECT ID, 設備名等, [労務費賃率] + [油等賃率] + [電気賃率] + [設備費賃率] + [建屋] + [土地] AS 賃率計 "
            "FROM t_賃率マスタ "
            "WHERE 工程分類 = '未選択' OR 工程分類 = '洗浄' "
            "ORDER BY ID;"
        )
        rows = cur.fetchall()
        kakou_24_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "rate_sum": _format_rate_sum_3(r[2]),
            }
            for r in rows
        ]

        cur.execute(
            "SELECT t_表面処理マスタ.ID, t_表面処理マスタ.表面処理名 FROM t_表面処理マスタ ORDER BY 並び順;"
        )
        rows = cur.fetchall()
        surface_options = [
            {"id": (r[0] if r[0] is not None else ""), "name": (r[1] if r[1] is not None else "")}
            for r in rows
        ]
        kakou_32_options = surface_options
        kakou_34_options = surface_options

        cur.execute(
            "SELECT ID, 設備名等, [労務費賃率] + [油等賃率] + [電気賃率] + [設備費賃率] + [建屋] + [土地] AS 賃率計 "
            "FROM t_賃率マスタ "
            "WHERE 工程分類 = '未選択' OR 工程分類 = '計量梱包' ORDER BY ID;"
        )
        rows = cur.fetchall()
        kakou_36_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "rate_sum": _format_rate_sum_3(r[2]),
            }
            for r in rows
        ]

        cur.execute("SELECT ID, 都道府県名, 地方ID FROM t_都道府県 ORDER BY ID;")
        rows = cur.fetchall()
        soryo_2_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "reg": (r[2] if len(r) > 2 and r[2] is not None else ""),
            }
            for r in rows
        ]

        cur.execute("SELECT ID, サイズ名, 大きさ, 重量 FROM t_運賃表サイズ ORDER BY ID;")
        rows = cur.fetchall()
        soryo_3_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "size": (r[2] if r[2] is not None else ""),
                "weight": (r[3] if r[3] is not None else ""),
            }
            for r in rows
        ]

        cur.execute("SELECT ID, 規格, 単価 FROM t_段ボール ORDER BY ID;")
        rows = cur.fetchall()
        soryo_8_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "price": (r[2] if r[2] is not None else ""),
            }
            for r in rows
        ]

        cur.execute("SELECT ID, トレー名, 収容数, 単価 FROM t_トレー ORDER BY ID;")
        rows = cur.fetchall()
        soryo_13_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "capacity": (r[2] if r[2] is not None else ""),
                "price": (r[3] if r[3] is not None else ""),
            }
            for r in rows
        ]

        # t_トレー: ID, トレー名, 材質, 収容数, 単価（表示はトレー名のみ／樹脂箱 soryo-29）
        cur.execute(
            "SELECT ID, トレー名, 材質, 収容数, 単価 FROM t_トレー ORDER BY ID;"
        )
        rows = cur.fetchall()
        soryo_29_options = [
            {
                "id": (r[0] if r[0] is not None else ""),
                "name": (r[1] if r[1] is not None else ""),
                "capacity": (r[3] if len(r) > 3 and r[3] is not None else ""),
                "price": (r[4] if len(r) > 4 and r[4] is not None else ""),
            }
            for r in rows
        ]

        def _rate_sum_by_equipment_name(equipment_name):
            cur.execute(
                "SELECT [労務費賃率] + [油等賃率] + [電気賃率] + [設備費賃率] + [建屋] + [土地] AS 賃率計 "
                "FROM t_賃率マスタ WHERE 設備名等 = ?",
                [equipment_name],
            )
            r = cur.fetchone()
            if not r or r[0] is None:
                return ""
            return _format_rate_sum_3(r[0])

        rate_default_map["blast"] = _rate_sum_by_equipment_name("ブラスト")
        rate_default_map["pressfit"] = _rate_sum_by_equipment_name("圧入")
        rate_default_map["pre_inspection"] = _rate_sum_by_equipment_name("処理前検査")
        rate_default_map["kensa_auto"] = _rate_sum_by_equipment_name("自動外観検査")
        rate_default_map["kensa_numeric"] = _rate_sum_by_equipment_name("数値")
        rate_default_map["kensa_visual"] = _rate_sum_by_equipment_name("目視")
        rate_default_map["kensa_microscope"] = _rate_sum_by_equipment_name("顕微鏡")
        rate_default_map["kensa_microgauge"] = _rate_sum_by_equipment_name("マイクロゲージ")
        rate_default_map["kensa_other"] = _rate_sum_by_equipment_name("その他")

        conn.close()
    except Exception:
        # エラー時はプルダウン空のまま
        pass

    if estimate_id:
        try:
            conn = get_connection()
            cur = conn.cursor()

            cur.execute(_est_calc_detail_sql(), [estimate_id])
            rows = cur.fetchall()

            if len(rows) == 1:
                row = rows[0]
                col_names = [c[0] for c in cur.description]
                rec = dict(zip(col_names, row))

                # 基本情報(t_原価見積履歴,t_原価見積情報)
                est_1 = _rec_str(rec, "品番")
                est_2 = _rec_str(rec, "品名")
                est_3 = _rec_str(rec, "客先名")
                est_4 = _rec_str(rec, "営業担当")
                est_5 = _rec_str(rec, "原価見積りID")
                est_6 = _rec_str(rec, "ロットID")
                est_7 = _rec_str(rec, "ロット数")
                est_8 = _rec_str(rec, "CT")
                est_9 = _rec_str(rec, "日産数")
                est_10 = _rec_str(rec, "日産数入力")
                est_11 = _rec_str(rec, "日数")
                est_12 = _rec_str(rec, "日数補正")

                # 材料費・真鍮（t_原価見積材料）
                mat_1 = _rec_str(rec, "材質径")
                mat_2 = _rec_str(rec, "鋼種")
                mat_3 = _rec_str(rec, "単重")
                mat_4 = _rec_str(rec, "形状")
                mat_5 = _rec_str(rec, "径")
                mat_6 = _rec_str(rec, "長さ")
                mat_7 = _rec_str(rec, "全長")
                mat_8 = _rec_str(rec, "突切り")
                mat_9 = _rec_str(rec, "取り数")
                mat_10 = _rec_str(rec, "取り数入力")
                mat_11 = _rec_str(rec, "材料費合計")
                mat_12 = _rec_str(rec, "比重")
                mat_13 = _rec_str(rec, "一本重")
                mat_14 = _rec_str(rec, "一個重")
                mat_15 = _rec_str(rec, "材料単価")
                mat_16 = _rec_str(rec, "材料費")
                mat_17 = format_yield_rate_display(rec.get("歩留り"))
                mat_18 = _rec_str(rec, "歩留り金額")
                mat_19 = _rec_str(rec, "材料費入力")

                # 真鍮詳細: t_原価見積真鍮 に行がない場合はチェック・入力をバインドしない
                if est_6:
                    cur.execute(
                        "SELECT COUNT(*) FROM t_原価見積真鍮 WHERE ロットID = ?",
                        [est_6],
                    )
                    _cnt = cur.fetchone()
                    shinchuu_has_row = bool(_cnt and _cnt[0] and _cnt[0] > 0)

                if shinchuu_has_row:
                    _r1 = _rec_str(rec, "RM区分")
                    shinchuu_r1 = _r1 if _r1 in ("1", "2") else "1"
                    _r2 = _rec_str(rec, "重量計算区分")
                    shinchuu_r2 = _r2 if _r2 in ("1", "2") else "1"
                    br_1 = _rec_str(rec, "素材単価")
                    br_2 = _rec_str(rec, "N社価格")
                    br_3 = _rec_str(rec, "建値")
                    br_4 = _rec_str(rec, "増値")
                    br_5 = _rec_str(rec, "真鍮単重")
                    br_6 = _rec_str(rec, "スクラップ重")
                    br_7 = _rec_str(rec, "スクラップベース")
                    br_8 = _rec_str(rec, "切粉回収率")
                    br_9 = _rec_str(rec, "スクラップ単価")
                    br_10 = _rec_str(rec, "真鍮材料費")

                # 加工費・管理費（t_原価見積加工管理）
                proc_1 = _rec_str(rec, "刃工具価格")
                proc_2 = _rec_str(rec, "材質")
                proc_3 = _rec_str(rec, "交換サイクル")
                proc_4 = _rec_str(rec, "刃工具原価")
                proc_5 = _rec_str(rec, "切削機械")
                proc_6 = _rec_str(rec, "切削機械サイクル")
                proc_7 = _rec_str(rec, "切削機械賃率")
                proc_8 = _rec_str(rec, "切削機械原価")
                proc_cb_1 = (_rec_str(rec, "ガイドブッシュ切替") == "-1")
                _set_level = _rec_str(rec, "セット難易度")
                proc_9 = _set_level if _set_level in ("1", "2", "3") else ""
                proc_10 = _rec_str(rec, "セット時間H")
                proc_11 = _rec_str(rec, "セット時間S")
                proc_12 = _rec_str(rec, "セット金額")
                proc_13 = _rec_str(rec, "セット原価")
                proc_cb_2 = (_rec_str(rec, "バレル切替") == "-1")
                proc_14 = _rec_str(rec, "バレル")
                proc_15 = _rec_str(rec, "バレルサイクル")
                proc_16 = _rec_str(rec, "バレル賃率")
                proc_17 = _rec_str(rec, "バレル原価")
                proc_cb_3 = (_rec_str(rec, "ブラスト切替") == "-1")
                proc_18 = _rec_str(rec, "ブラストサイクル")
                proc_19 = _rec_str(rec, "ブラスト賃率")
                proc_20 = _rec_str(rec, "ブラスト原価")
                proc_cb_4 = (_rec_str(rec, "圧入切替") == "-1")
                proc_21 = _rec_str(rec, "圧入サイクル")
                proc_22 = _rec_str(rec, "圧入賃率")
                proc_23 = _rec_str(rec, "圧入原価")
                proc_24 = _rec_str(rec, "洗浄")
                proc_25 = _rec_str(rec, "洗浄サイクル")
                proc_26 = _rec_str(rec, "洗浄賃率")
                proc_27 = _rec_str(rec, "洗浄原価")
                proc_cb_5 = (_rec_str(rec, "処理前検査切替") == "-1")
                proc_28 = _rec_str(rec, "処理前検査サイクル")
                proc_29 = _rec_str(rec, "処理前検査賃率")
                proc_30 = _rec_str(rec, "処理前検査原価")
                proc_cb_6 = (_rec_str(rec, "表面処理切替") == "-1")
                proc_31 = est_7 or _rec_str(rec, "ロット数")
                proc_32 = _rec_str(rec, "表面処理名ID")
                proc_33 = _rec_str(rec, "表面処理原価")
                proc_34 = _rec_str(rec, "表面処理名ID2")
                proc_35 = _rec_str(rec, "表面処理原価2")
                proc_36 = _rec_str(rec, "計量梱包")
                proc_37 = _rec_str(rec, "計量梱包サイクル")
                proc_38 = _rec_str(rec, "計量梱包賃率")
                proc_39 = _rec_str(rec, "計量梱包原価")

                # 検査（t_原価見積加工管理）
                kensa_cb_1 = (_rec_str(rec, "自動外観検査切替") == "-1")
                kensa_1 = _rec_str(rec, "自動外観検査サイクル")
                kensa_2 = _rec_str(rec, "自動外観検査賃率")
                kensa_3 = _rec_str(rec, "自動外観検査原価")
                kensa_cb_2 = (_rec_str(rec, "数値切替") == "-1")
                kensa_4 = _rec_str(rec, "数値サイクル")
                kensa_5 = _rec_str(rec, "数値賃率")
                kensa_6 = _rec_str(rec, "数値原価")
                kensa_cb_3 = (_rec_str(rec, "目視切替") == "-1")
                kensa_7 = _rec_str(rec, "目視サイクル")
                kensa_8 = _rec_str(rec, "目視賃率")
                kensa_9 = _rec_str(rec, "目視原価")
                kensa_cb_4 = (_rec_str(rec, "顕微鏡切替") == "-1")
                kensa_10 = _rec_str(rec, "顕微鏡サイクル")
                kensa_11 = _rec_str(rec, "顕微鏡賃率")
                kensa_12 = _rec_str(rec, "顕微鏡原価")
                kensa_cb_5 = (_rec_str(rec, "マイクロゲージ切替") == "-1")
                kensa_13 = _rec_str(rec, "マイクロゲージサイクル")
                kensa_14 = _rec_str(rec, "マイクロゲージ賃率")
                kensa_15 = _rec_str(rec, "マイクロゲージ原価")
                kensa_cb_6 = (_rec_str(rec, "その他切替") == "-1")
                kensa_16 = _rec_str(rec, "その他検査名")
                kensa_17 = _rec_str(rec, "その他サイクル")
                kensa_18 = _rec_str(rec, "その他賃率")
                kensa_19 = _rec_str(rec, "その他原価")
                kensa_20 = _rec_str(rec, "その他検査名2")
                kensa_21 = _rec_str(rec, "その他サイクル2")
                kensa_22 = _rec_str(rec, "その他原価2")
                kensa_23 = _rec_str(rec, "その他検査名3")
                kensa_24 = _rec_str(rec, "その他サイクル3")
                kensa_25 = _rec_str(rec, "その他原価3")
                kensa_26 = _rec_str(rec, "その他検査名4")
                kensa_27 = _rec_str(rec, "その他サイクル4")
                kensa_28 = _rec_str(rec, "その他原価4")
                kensa_29 = _rec_str(rec, "その他検査名5")
                kensa_30 = _rec_str(rec, "その他サイクル5")
                kensa_31 = _rec_str(rec, "その他原価5")

                # 送料・梱包（t_原価見積送料）
                _box = _rec_str(rec, "使用箱切替")
                est_soryo_box = _box if _box in ("1", "2") else "1"
                soryo_1 = _rec_str(rec, "D入数")
                soryo_2 = _rec_str(rec, "D納入先")
                soryo_3 = _rec_str(rec, "D梱包サイズ")
                soryo_4 = _rec_str(rec, "D箱サイズ")
                soryo_5 = _rec_str(rec, "D重量")
                soryo_6 = _rec_str(rec, "D送料")
                soryo_7 = _rec_str(rec, "D運賃単価")
                soryo_8 = _rec_str(rec, "D箱規格")
                soryo_9 = _rec_str(rec, "D箱価格")
                soryo_10 = _rec_str(rec, "D箱入数")
                soryo_11 = _rec_str(rec, "D一箱重量")
                soryo_12 = _rec_str(rec, "D箱単価")
                soryo_13 = _rec_str(rec, "D使用トレー")
                soryo_14 = _rec_str(rec, "D収容数")
                soryo_15 = _rec_str(rec, "Dトレー価格")
                soryo_16 = _rec_str(rec, "D天井フタ数")
                soryo_17 = _rec_str(rec, "D必要トレー数")
                soryo_18 = _rec_str(rec, "Dトレー合計額")
                soryo_19 = _rec_str(rec, "Dトレー単価")
                soryo_36 = _rec_str(rec, "納入日数")
                soryo_37 = _rec_str(rec, "納入数日")
                soryo_38 = _rec_str(rec, "日当重量")
                soryo_39 = _rec_str(rec, "日当送料単価")
                soryo_20 = _rec_str(rec, "J単価パレット")
                soryo_21 = _rec_str(rec, "J搭載箱数パレット")
                soryo_22 = _rec_str(rec, "J搭載数パレット")
                soryo_23 = _rec_str(rec, "J運賃単価パレット")
                soryo_24 = _rec_str(rec, "J箱負担")
                soryo_25 = _rec_str(rec, "J箱規格")
                soryo_26 = _rec_str(rec, "J箱価格")
                soryo_27 = _rec_str(rec, "J箱入数")
                soryo_28 = _rec_str(rec, "J箱単価")
                soryo_29 = _rec_str(rec, "J使用トレー")
                soryo_30 = _rec_str(rec, "J収容数")
                soryo_31 = _rec_str(rec, "Jトレー価格")
                soryo_32 = _rec_str(rec, "J天井フタ数")
                soryo_33 = _rec_str(rec, "J必要トレー数")
                soryo_34 = _rec_str(rec, "Jトレー合計額")
                soryo_35 = _rec_str(rec, "Jトレー単価")

                # 単価計算等（t_原価見積計算チャージ）
                f_in_1 = _rec_str(rec, "加工管理合計社内")
                f_in_2 = _rec_str(rec, "管理費率社内")
                f_in_3 = _rec_str(rec, "管理費社内")
                f_out_1 = _rec_str(rec, "表面処理費")
                f_out_2 = _rec_str(rec, "管理費率社外")
                f_out_3 = _rec_str(rec, "管理費社外")
                f_out_4 = _rec_str(rec, "表面処理費2")
                f_out_5 = _rec_str(rec, "管理費率社外2")
                f_out_6 = _rec_str(rec, "管理費社外2")
                f_in_b1 = _rec_str(rec, "原価合計")
                f_in_b2 = _rec_str(rec, "粗利率")
                f_in_b3 = _rec_str(rec, "粗利")
                f_in_b4 = _rec_str(rec, "見積単価")
                f_ch_1 = _rec_str(rec, "チャージ材料費")
                f_ch_2 = _rec_str(rec, "チャージ刃工具費")
                f_ch_3 = _rec_str(rec, "チャージ社外管理費")
                f_ch_4 = _rec_str(rec, "チャージ表面処理費")
                f_ch_5 = _rec_str(rec, "チャージ検査費")
                f_ch_6 = _rec_str(rec, "チャージ梱包費")
                f_ch_7 = _rec_str(rec, "機械チャージ")
                f_ch_8 = _rec_str(rec, "検査チャージ")
                f_ch_9 = _rec_str(rec, "梱包チャージ")
                f_ch_10 = _rec_str(rec, "チャージ金額")

            elif len(rows) == 0:
                # 新規登録直後など t_原価見積情報が無いと詳細クエリは 0 行になる。
                # 履歴＋マスタから品番・品名・客先名・営業担当・原価見積りIDのみ反映する。
                cur.execute(
                    "SELECT "
                    "t_原価見積履歴.品番, "
                    "t_原価見積履歴.品名, "
                    "t_客先マスタ.客先名, "
                    "t_営業マスタ.営業担当, "
                    "t_原価見積履歴.原価見積りID "
                    "FROM (t_原価見積履歴 "
                    "LEFT JOIN t_営業マスタ ON t_原価見積履歴.営業ID = t_営業マスタ.コード) "
                    "LEFT JOIN t_客先マスタ ON t_原価見積履歴.客先コード = t_客先マスタ.コード "
                    "WHERE t_原価見積履歴.原価見積りID = ?",
                    [estimate_id],
                )
                _kihon = cur.fetchone()
                if _kihon:
                    est_1 = "" if _kihon[0] is None else str(_kihon[0])
                    est_2 = "" if _kihon[1] is None else str(_kihon[1])
                    est_3 = "" if _kihon[2] is None else str(_kihon[2])
                    est_4 = "" if _kihon[3] is None else str(_kihon[3])
                    est_5 = "" if _kihon[4] is None else str(_kihon[4])

            # ロット一覧用クエリ（ロットID, ロット数, 見積単価 を表示）
            lot_sql = (
                "SELECT t_原価見積情報.ロットID, t_原価見積情報.原価見積りID, "
                "t_原価見積情報.使用フラグ, t_原価見積情報.ロット数, "
                "t_原価見積計算チャージ.ロットID, t_原価見積計算チャージ.見積単価 "
                "FROM t_原価見積情報 "
                "LEFT JOIN t_原価見積計算チャージ "
                "ON t_原価見積情報.ロットID = t_原価見積計算チャージ.ロットID "
                "WHERE t_原価見積情報.原価見積りID = ? "
                "ORDER BY t_原価見積情報.使用フラグ DESC, "
                "t_原価見積情報.ロット数 ASC, t_原価見積情報.ロットID ASC"
            )
            cur.execute(lot_sql, [estimate_id])
            lot_result = cur.fetchall()
            # 表示用: ロットID(0), ロット数(3), 見積単価(5)
            for r in lot_result:
                v0 = r[0] if r[0] is not None else ""
                v3 = r[3] if r[3] is not None else ""
                v5 = r[5] if r[5] is not None else ""
                lot_rows.append({"ロットID": v0, "ロット数": v3, "見積単価": v5})

            initial_cost_rows.extend(_initial_cost_rows_for_api(cur, estimate_id))

            conn.close()

        except Exception:
            # エラー時は空のまま
            pass

    br_8_display = format_cutting_recovery_rate_display(br_8)

    # 送料・梱包: 梱包サイズが未設定なら ID=4 を既定値にし、箱サイズ/重量も合わせる
    if not (soryo_3 or "").strip():
        default_pack = next(
            (opt for opt in soryo_3_options if str(opt.get("id", "")).strip() == "4"),
            None,
        )
        if default_pack is not None:
            soryo_3 = "4"
            soryo_4 = str(default_pack.get("size", "") or "")
            soryo_5 = str(default_pack.get("weight", "") or "")

    return {
        "ok": True,
        "est_1": est_1,
        "est_2": est_2,
        "est_3": est_3,
        "est_4": est_4,
        "est_5": est_5,
        "est_6": est_6,
        "est_7": est_7,
        "est_8": est_8,
        "est_9": est_9,
        "est_10": est_10,
        "est_11": est_11,
        "est_12": est_12,
        "mat_1": mat_1,
        "mat_2": mat_2,
        "mat_3": mat_3,
        "mat_4": mat_4,
        "mat_5": mat_5,
        "mat_6": mat_6,
        "mat_7": mat_7,
        "mat_8": mat_8,
        "mat_9": mat_9,
        "mat_10": mat_10,
        "mat_11": mat_11,
        "mat_12": mat_12,
        "mat_13": mat_13,
        "mat_14": mat_14,
        "mat_15": mat_15,
        "mat_16": mat_16,
        "mat_17": mat_17,
        "mat_18": mat_18,
        "mat_19": mat_19,
        "shinchuu_has_row": shinchuu_has_row,
        "shinchuu_r1": shinchuu_r1,
        "shinchuu_r2": shinchuu_r2,
        "br_1": br_1,
        "br_2": br_2,
        "br_3": br_3,
        "br_4": br_4,
        "br_5": br_5,
        "br_6": br_6,
        "br_7": br_7,
        "br_8": br_8,
        "br_8_display": br_8_display,
        "br_9": br_9,
        "br_10": br_10,
        "proc_1": proc_1,
        "proc_2": proc_2,
        "proc_3": proc_3,
        "proc_4": proc_4,
        "proc_5": proc_5,
        "proc_6": proc_6,
        "proc_7": proc_7,
        "proc_8": proc_8,
        "proc_9": proc_9,
        "proc_10": proc_10,
        "proc_11": proc_11,
        "proc_12": proc_12,
        "proc_13": proc_13,
        "proc_14": proc_14,
        "proc_15": proc_15,
        "proc_16": proc_16,
        "proc_17": proc_17,
        "proc_18": proc_18,
        "proc_19": proc_19,
        "proc_20": proc_20,
        "proc_21": proc_21,
        "proc_22": proc_22,
        "proc_23": proc_23,
        "proc_24": proc_24,
        "proc_25": proc_25,
        "proc_26": proc_26,
        "proc_27": proc_27,
        "proc_28": proc_28,
        "proc_29": proc_29,
        "proc_30": proc_30,
        "proc_31": proc_31,
        "proc_32": proc_32,
        "proc_33": proc_33,
        "proc_34": proc_34,
        "proc_35": proc_35,
        "proc_36": proc_36,
        "proc_37": proc_37,
        "proc_38": proc_38,
        "proc_39": proc_39,
        "proc_cb_1": proc_cb_1,
        "proc_cb_2": proc_cb_2,
        "proc_cb_3": proc_cb_3,
        "proc_cb_4": proc_cb_4,
        "proc_cb_5": proc_cb_5,
        "proc_cb_6": proc_cb_6,
        "kensa_1": kensa_1,
        "kensa_2": kensa_2,
        "kensa_3": kensa_3,
        "kensa_4": kensa_4,
        "kensa_5": kensa_5,
        "kensa_6": kensa_6,
        "kensa_7": kensa_7,
        "kensa_8": kensa_8,
        "kensa_9": kensa_9,
        "kensa_10": kensa_10,
        "kensa_11": kensa_11,
        "kensa_12": kensa_12,
        "kensa_13": kensa_13,
        "kensa_14": kensa_14,
        "kensa_15": kensa_15,
        "kensa_16": kensa_16,
        "kensa_17": kensa_17,
        "kensa_18": kensa_18,
        "kensa_19": kensa_19,
        "kensa_20": kensa_20,
        "kensa_21": kensa_21,
        "kensa_22": kensa_22,
        "kensa_23": kensa_23,
        "kensa_24": kensa_24,
        "kensa_25": kensa_25,
        "kensa_26": kensa_26,
        "kensa_27": kensa_27,
        "kensa_28": kensa_28,
        "kensa_29": kensa_29,
        "kensa_30": kensa_30,
        "kensa_31": kensa_31,
        "kensa_cb_1": kensa_cb_1,
        "kensa_cb_2": kensa_cb_2,
        "kensa_cb_3": kensa_cb_3,
        "kensa_cb_4": kensa_cb_4,
        "kensa_cb_5": kensa_cb_5,
        "kensa_cb_6": kensa_cb_6,
        "est_soryo_box": est_soryo_box,
        "soryo_1": soryo_1,
        "soryo_2": soryo_2,
        "soryo_3": soryo_3,
        "soryo_4": soryo_4,
        "soryo_5": soryo_5,
        "soryo_6": soryo_6,
        "soryo_7": soryo_7,
        "soryo_8": soryo_8,
        "soryo_9": soryo_9,
        "soryo_10": soryo_10,
        "soryo_11": soryo_11,
        "soryo_12": soryo_12,
        "soryo_13": soryo_13,
        "soryo_14": soryo_14,
        "soryo_15": soryo_15,
        "soryo_16": soryo_16,
        "soryo_17": soryo_17,
        "soryo_18": soryo_18,
        "soryo_19": soryo_19,
        "soryo_20": soryo_20,
        "soryo_21": soryo_21,
        "soryo_22": soryo_22,
        "soryo_23": soryo_23,
        "soryo_24": soryo_24,
        "soryo_25": soryo_25,
        "soryo_26": soryo_26,
        "soryo_27": soryo_27,
        "soryo_28": soryo_28,
        "soryo_29": soryo_29,
        "soryo_30": soryo_30,
        "soryo_31": soryo_31,
        "soryo_32": soryo_32,
        "soryo_33": soryo_33,
        "soryo_34": soryo_34,
        "soryo_35": soryo_35,
        "soryo_36": soryo_36,
        "soryo_37": soryo_37,
        "soryo_38": soryo_38,
        "soryo_39": soryo_39,
        "f_in_1": f_in_1,
        "f_in_2": f_in_2,
        "f_in_3": f_in_3,
        "f_out_1": f_out_1,
        "f_out_2": f_out_2,
        "f_out_3": f_out_3,
        "f_out_4": f_out_4,
        "f_out_5": f_out_5,
        "f_out_6": f_out_6,
        "f_in_b1": f_in_b1,
        "f_in_b2": f_in_b2,
        "f_in_b3": f_in_b3,
        "f_in_b4": f_in_b4,
        "f_ch_1": f_ch_1,
        "f_ch_2": f_ch_2,
        "f_ch_3": f_ch_3,
        "f_ch_4": f_ch_4,
        "f_ch_5": f_ch_5,
        "f_ch_6": f_ch_6,
        "f_ch_7": f_ch_7,
        "f_ch_8": f_ch_8,
        "f_ch_9": f_ch_9,
        "f_ch_10": f_ch_10,
        "lot_rows": lot_rows,
        "zairyo_2_options": zairyo_2_options,
        "kakou_2_options": kakou_2_options,
        "kakou_5_options": kakou_5_options,
        "kakou_14_options": kakou_14_options,
        "kakou_24_options": kakou_24_options,
        "kakou_32_options": kakou_32_options,
        "kakou_34_options": kakou_34_options,
        "kakou_36_options": kakou_36_options,
        "soryo_2_options": soryo_2_options,
        "soryo_3_options": soryo_3_options,
        "soryo_8_options": soryo_8_options,
        "soryo_13_options": soryo_13_options,
        "soryo_29_options": soryo_29_options,
        "rate_default_map": rate_default_map,
        "initial_cost_rows": initial_cost_rows,
    }


def _est_calc_detail_sql():
    """est_calc で使用する詳細取得クエリ（1件用）"""
    return (
        "SELECT * FROM ((((((("
        "t_原価見積履歴 "
        "LEFT JOIN t_原価見積情報 "
        "ON t_原価見積履歴.原価見積りID = t_原価見積情報.原価見積りID) "
        "LEFT JOIN t_原価見積材料 "
        "ON t_原価見積情報.ロットID = t_原価見積材料.ロットID) "
        "LEFT JOIN t_原価見積真鍮 "
        "ON t_原価見積情報.ロットID = t_原価見積真鍮.ロットID) "
        "LEFT JOIN t_原価見積加工管理 "
        "ON t_原価見積情報.ロットID = t_原価見積加工管理.ロットID) "
        "LEFT JOIN t_原価見積送料 "
        "ON t_原価見積情報.ロットID = t_原価見積送料.ロットID) "
        "LEFT JOIN t_原価見積計算チャージ "
        "ON t_原価見積情報.ロットID = t_原価見積計算チャージ.ロットID) "
        "LEFT JOIN t_営業マスタ "
        "ON t_原価見積履歴.営業ID = t_営業マスタ.コード) "
        "LEFT JOIN t_客先マスタ "
        "ON t_原価見積履歴.客先コード = t_客先マスタ.コード "
        "WHERE t_原価見積履歴.原価見積りID = ? "
        "AND t_原価見積情報.使用フラグ = 'Y';"
    )


def _json_safe_cell_value(v):
    """pyodbc のセル値を jsonify 可能な型に変換する"""
    if v is None:
        return None
    from decimal import Decimal
    from datetime import datetime, date, time

    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):
        return v.isoformat()
    if isinstance(v, (bytes, bytearray)):
        return v.decode("utf-8", errors="replace")
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def _est_calc_detail_row_to_jsonable_dict(row, col_names):
    """詳細クエリ1行を JSON 用 dict に変換（列名は cur.description と同一）"""
    return {k: _json_safe_cell_value(val) for k, val in zip(col_names, row)}


def _sanitize_filename_part(value, max_len=None):
    """Windows で使えない文字を除去し、必要なら長さを制限する"""
    s = (value or "").strip()
    s = s.replace("\r", " ").replace("\n", " ")
    invalid = '<>:"/\\|?*'
    for ch in invalid:
        s = s.replace(ch, "_")
    if max_len is not None:
        s = s[:max_len]
    return s


def api_est_calc_set_lot(payload=None):
    """ロットIDクリック時: 使用フラグを更新し、扱うデータを変更"""
    data = payload or {}
    estimate_id = (data.get('estimate_id') or '').strip()  # 原価見積りID (est_1)
    current_lot_id = (data.get('current_lot_id') or '').strip()  # est_2 で表示中のロットID
    click_lot_id = (data.get('lot_id') or '').strip()  # クリックした行のロットID

    if not estimate_id or not click_lot_id:
        return {"error": "estimate_id と lot_id が必要です"}

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()

        # 使用フラグを空にする: WHERE 原価見積りID = ? の ? = est_2で表示しているロットID
        cur.execute(
            "UPDATE t_原価見積情報 SET 使用フラグ = '' WHERE 原価見積りID = ?",
            [estimate_id],
        )

        # クリックしたロットIDの行だけ 使用フラグ = 'Y' にする
        cur.execute(
            "UPDATE t_原価見積情報 SET 使用フラグ = 'Y' WHERE ロットID = ?",
            [click_lot_id],
        )

        conn.commit()

        # 使用フラグ更新後、est_calc と同じ詳細取得クエリ（原価見積りID = estimate_id）
        cur.execute(_est_calc_detail_sql(), [estimate_id])
        detail_rows = cur.fetchall()
        col_names = [c[0] for c in cur.description] if cur.description else []
        detail = None
        if detail_rows:
            detail = _est_calc_detail_row_to_jsonable_dict(detail_rows[0], col_names)

        return {"ok": True, "estimate_id": estimate_id, "detail": detail}
    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass

def api_est_calc_clear_usage_flag(payload=None):
    """見積り追加(1段目)用: 指定見積りIDの使用フラグを全て空にする"""
    data = payload or {}
    estimate_id = (data.get("estimate_id") or "").strip()
    if not estimate_id:
        return {"error": "estimate_id が必要です"}

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "UPDATE t_原価見積情報 SET 使用フラグ = '' WHERE 原価見積りID = ?",
            [estimate_id],
        )
        conn.commit()
        return {"ok": True, "estimate_id": estimate_id}
    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def api_est_calc_add_estimate_lot(payload=None):
    """見積り追加(2段目)用: 新規ロットを作成し、コピー有無で各テーブルへ追加"""
    data = payload or {}
    estimate_id = (data.get("estimate_id") or "").strip()
    copy_data = bool(data.get("copy_data"))

    if not estimate_id:
        return {"error": "estimate_id が必要です"}

    # payload には buildSavePayload() の値（各種 est_/mat_/br_/proc_/kensa_/soryo_ など）が入る想定
    # ただし lot_id は新規ロット作成後に上書きして使う
    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()

        # t_原価見積情報（新規ロット）を追加（使用フラグのみ有効）
        cur.execute(
            "INSERT INTO t_原価見積情報 (原価見積りID, 使用フラグ) VALUES (?, 'Y')",
            [estimate_id],
        )

        # 新しいロットID取得
        cur.execute("SELECT MAX(ロットID) FROM t_原価見積情報;")
        row = cur.fetchone()
        lot_id = row[0] if row else None
        if lot_id is None:
            raise Exception("新しいロットIDの取得に失敗しました")

        if copy_data:
            if "br_8" in data:
                data["br_8"] = normalize_cutting_recovery_rate_for_db(data.get("br_8"))
            if "mat_17" in data:
                data["mat_17"] = normalize_yield_rate_for_db(data.get("mat_17"))

            nz_zero_keys = {
                "est_7", "est_8", "est_9", "est_10", "est_11", "est_12",
                "mat_2", "mat_3", "mat_5", "mat_6", "mat_7", "mat_8", "mat_9", "mat_10",
                "mat_11", "mat_12", "mat_13", "mat_14", "mat_15", "mat_16", "mat_17", "mat_18", "mat_19",
                "br_1", "br_2", "br_3", "br_4", "br_5", "br_6", "br_7", "br_8", "br_9", "br_10",
                "proc_1", "proc_2", "proc_3", "proc_4", "proc_5", "proc_6", "proc_7", "proc_8",
                "proc_9", "proc_10", "proc_11", "proc_12", "proc_13", "proc_14", "proc_15", "proc_16", "proc_17",
                "proc_18", "proc_19", "proc_20", "proc_21", "proc_22", "proc_23", "proc_24", "proc_25", "proc_26", "proc_27",
                "proc_28", "proc_29", "proc_30", "proc_32", "proc_33", "proc_34", "proc_35", "proc_36", "proc_37", "proc_38", "proc_39",
                "kensa_1", "kensa_2", "kensa_3", "kensa_4", "kensa_5", "kensa_6", "kensa_7", "kensa_8", "kensa_9",
                "kensa_10", "kensa_11", "kensa_12", "kensa_13", "kensa_14", "kensa_15",
                "kensa_17", "kensa_18", "kensa_19", "kensa_21", "kensa_22", "kensa_24", "kensa_25", "kensa_27", "kensa_28", "kensa_30", "kensa_31",
                "soryo_1", "soryo_2", "soryo_3", "soryo_6", "soryo_7",
                "soryo_9", "soryo_10", "soryo_11", "soryo_12", "soryo_13", "soryo_14", "soryo_15", "soryo_16", "soryo_17", "soryo_18", "soryo_19",
                "soryo_36", "soryo_37", "soryo_38", "soryo_39", "soryo_20", "soryo_21", "soryo_22", "soryo_23",
                "soryo_26", "soryo_27", "soryo_28", "soryo_29", "soryo_30", "soryo_31", "soryo_32", "soryo_33", "soryo_34", "soryo_35",
                "f_in_1", "f_in_2", "f_in_3", "f_out_1", "f_out_2", "f_out_3", "f_out_4", "f_out_5", "f_out_6",
                "f_in_b1", "f_in_b2", "f_in_b3", "f_in_b4", "f_ch_1", "f_ch_2", "f_ch_3", "f_ch_4", "f_ch_5", "f_ch_6", "f_ch_7", "f_ch_8", "f_ch_9", "f_ch_10",
            }

            def _val(key):
                v = data.get(key)
                if key in nz_zero_keys:
                    if v is None:
                        return "0"
                    s = str(v).strip()
                    if s == "":
                        return "0"
                    return s
                if v is None:
                    return ""
                return str(v)

            def _run_upsert(cur, table_name, pairs):
                cur.execute(f"SELECT COUNT(*) FROM {table_name} WHERE [ロットID] = ?", [lot_id])
                cnt_row = cur.fetchone()
                exists = bool(cnt_row and cnt_row[0] and cnt_row[0] > 0)

                if exists:
                    for col, key in pairs:
                        sql = f"UPDATE {table_name} SET [{col}] = ? WHERE [ロットID] = ?"
                        params = [_val(key), lot_id]
                        cur.execute(sql, params)
                else:
                    cols = "[ロットID], " + ", ".join([f"[{col}]" for col, _ in pairs])
                    placeholders = ", ".join(["?"] * (len(pairs) + 1))
                    sql = f"INSERT INTO {table_name} ({cols} VALUES ({placeholders})"
                    params = [lot_id] + [_val(k) for _, k in pairs]
                    cur.execute(sql, params)

            # t_原価見積情報
            _run_upsert(cur, "t_原価見積情報", [
                ("ロット数", "est_7"),
                ("CT", "est_8"),
                ("日産数", "est_9"),
                ("日産数入力", "est_10"),
                ("日数補正", "est_12"),
                ("日数", "est_11"),
            ])

            # t_原価見積材料
            _run_upsert(cur, "t_原価見積材料", [
                ("材質径", "mat_1"),
                ("鋼種", "mat_2"),
                ("単重", "mat_3"),
                ("形状", "mat_4"),
                ("径", "mat_5"),
                ("長さ", "mat_6"),
                ("全長", "mat_7"),
                ("突切り", "mat_8"),
                ("取り数", "mat_9"),
                ("取り数入力", "mat_10"),
                ("材料費合計", "mat_11"),
                ("比重", "mat_12"),
                ("一本重", "mat_13"),
                ("一個重", "mat_14"),
                ("材料単価", "mat_15"),
                ("材料費", "mat_16"),
                ("歩留り", "mat_17"),
                ("歩留り金額", "mat_18"),
                ("材料費入力", "mat_19"),
            ])

            # t_原価見積真鍮（有効化チェック時のみ）
            if bool(data.get("br_cb_1")):
                _run_upsert(cur, "t_原価見積真鍮", [
                    ("RM区分", "shinchuu_r1"),
                    ("重量計算区分", "shinchuu_r2"),
                    ("素材単価", "br_1"),
                    ("N社価格", "br_2"),
                    ("建値", "br_3"),
                    ("増値", "br_4"),
                    ("真鍮単重", "br_5"),
                    ("スクラップ重", "br_6"),
                    ("スクラップベース", "br_7"),
                    ("切粉回収率", "br_8"),
                    ("スクラップ単価", "br_9"),
                    ("真鍮材料費", "br_10"),
                ])

            # t_原価見積加工管理
            _run_upsert(cur, "t_原価見積加工管理", [
                ("刃工具価格", "proc_1"),
                ("材質", "proc_2"),
                ("交換サイクル", "proc_3"),
                ("刃工具原価", "proc_4"),
                ("切削機械", "proc_5"),
                ("切削機械サイクル", "proc_6"),
                ("切削機械賃率", "proc_7"),
                ("切削機械原価", "proc_8"),
                ("ガイドブッシュ切替", "proc_cb_1"),
                ("セット難易度", "proc_9"),
                ("セット時間H", "proc_10"),
                ("セット時間S", "proc_11"),
                ("セット金額", "proc_12"),
                ("セット原価", "proc_13"),
                ("バレル切替", "proc_cb_2"),
                ("バレル", "proc_14"),
                ("バレルサイクル", "proc_15"),
                ("バレル賃率", "proc_16"),
                ("バレル原価", "proc_17"),
                ("ブラスト切替", "proc_cb_3"),
                ("ブラストサイクル", "proc_18"),
                ("ブラスト賃率", "proc_19"),
                ("ブラスト原価", "proc_20"),
                ("圧入切替", "proc_cb_4"),
                ("圧入サイクル", "proc_21"),
                ("圧入賃率", "proc_22"),
                ("圧入原価", "proc_23"),
                ("洗浄", "proc_24"),
                ("洗浄サイクル", "proc_25"),
                ("洗浄賃率", "proc_26"),
                ("洗浄原価", "proc_27"),
                ("処理前検査切替", "proc_cb_5"),
                ("処理前検査サイクル", "proc_28"),
                ("処理前検査賃率", "proc_29"),
                ("処理前検査原価", "proc_30"),
                ("表面処理切替", "proc_cb_6"),
                ("表面処理名ID", "proc_32"),
                ("表面処理原価", "proc_33"),
                ("表面処理名ID2", "proc_34"),
                ("表面処理原価2", "proc_35"),
                ("計量梱包", "proc_36"),
                ("計量梱包サイクル", "proc_37"),
                ("計量梱包賃率", "proc_38"),
                ("計量梱包原価", "proc_39"),
                ("自動外観検査切替", "kensa_cb_1"),
                ("自動外観検査サイクル", "kensa_1"),
                ("自動外観検査賃率", "kensa_2"),
                ("自動外観検査原価", "kensa_3"),
                ("数値切替", "kensa_cb_2"),
                ("数値サイクル", "kensa_4"),
                ("数値賃率", "kensa_5"),
                ("数値原価", "kensa_6"),
                ("目視切替", "kensa_cb_3"),
                ("目視サイクル", "kensa_7"),
                ("目視賃率", "kensa_8"),
                ("目視原価", "kensa_9"),
                ("顕微鏡切替", "kensa_cb_4"),
                ("顕微鏡サイクル", "kensa_10"),
                ("顕微鏡賃率", "kensa_11"),
                ("顕微鏡原価", "kensa_12"),
                ("マイクロゲージ切替", "kensa_cb_5"),
                ("マイクロゲージサイクル", "kensa_13"),
                ("マイクロゲージ賃率", "kensa_14"),
                ("マイクロゲージ原価", "kensa_15"),
                ("その他切替", "kensa_cb_6"),
                ("その他検査名", "kensa_16"),
                ("その他サイクル", "kensa_17"),
                ("その他賃率", "kensa_18"),
                ("その他原価", "kensa_19"),
                ("その他検査名2", "kensa_20"),
                ("その他サイクル2", "kensa_21"),
                ("その他原価2", "kensa_22"),
                ("その他検査名3", "kensa_23"),
                ("その他サイクル3", "kensa_24"),
                ("その他原価3", "kensa_25"),
                ("その他検査名4", "kensa_26"),
                ("その他サイクル4", "kensa_27"),
                ("その他原価4", "kensa_28"),
                ("その他検査名5", "kensa_29"),
                ("その他サイクル5", "kensa_30"),
                ("その他原価5", "kensa_31"),
            ])

            # t_原価見積送料
            _run_upsert(cur, "t_原価見積送料", [
                ("使用箱切替", "est_soryo_box"),
                ("D入数", "soryo_1"),
                ("D納入先", "soryo_2"),
                ("D梱包サイズ", "soryo_3"),
                ("D箱サイズ", "soryo_4"),
                ("D重量", "soryo_5"),
                ("D送料", "soryo_6"),
                ("D運賃単価", "soryo_7"),
                ("D箱規格", "soryo_8"),
                ("D箱価格", "soryo_9"),
                ("D箱入数", "soryo_10"),
                ("D一箱重量", "soryo_11"),
                ("D箱単価", "soryo_12"),
                ("D使用トレー", "soryo_13"),
                ("D収容数", "soryo_14"),
                ("Dトレー価格", "soryo_15"),
                ("D天井フタ数", "soryo_16"),
                ("D必要トレー数", "soryo_17"),
                ("Dトレー合計額", "soryo_18"),
                ("Dトレー単価", "soryo_19"),
                ("納入日数", "soryo_36"),
                ("納入数日", "soryo_37"),
                ("日当重量", "soryo_38"),
                ("日当送料単価", "soryo_39"),
                ("J単価パレット", "soryo_20"),
                ("J搭載箱数パレット", "soryo_21"),
                ("J搭載数パレット", "soryo_22"),
                ("J運賃単価パレット", "soryo_23"),
                ("J箱負担", "soryo_24"),
                ("J箱規格", "soryo_25"),
                ("J箱価格", "soryo_26"),
                ("J箱入数", "soryo_27"),
                ("J箱単価", "soryo_28"),
                ("J使用トレー", "soryo_29"),
                ("J収容数", "soryo_30"),
                ("Jトレー価格", "soryo_31"),
                ("J天井フタ数", "soryo_32"),
                ("J必要トレー数", "soryo_33"),
                ("Jトレー合計額", "soryo_34"),
                ("Jトレー単価", "soryo_35"),
            ])

            # t_原価見積計算チャージ
            _run_upsert(cur, "t_原価見積計算チャージ", [
                ("加工管理合計社内", "f_in_1"),
                ("管理費率社内", "f_in_2"),
                ("管理費社内", "f_in_3"),
                ("表面処理費", "f_out_1"),
                ("管理費率社外", "f_out_2"),
                ("管理費社外", "f_out_3"),
                ("表面処理費2", "f_out_4"),
                ("管理費率社外2", "f_out_5"),
                ("管理費社外2", "f_out_6"),
                ("原価合計", "f_in_b1"),
                ("粗利率", "f_in_b2"),
                ("粗利", "f_in_b3"),
                ("見積単価", "f_in_b4"),
                ("チャージ材料費", "f_ch_1"),
                ("チャージ刃工具費", "f_ch_2"),
                ("チャージ社外管理費", "f_ch_3"),
                ("チャージ表面処理費", "f_ch_4"),
                ("チャージ検査費", "f_ch_5"),
                ("チャージ梱包費", "f_ch_6"),
                ("機械チャージ", "f_ch_7"),
                ("検査チャージ", "f_ch_8"),
                ("梱包チャージ", "f_ch_9"),
                ("チャージ金額", "f_ch_10"),
            ])

        else:
            # 空データ追加: 材料/加工管理/送料/計算チャージにロットIDだけ入れる（真鍮は追加しない）
            def _ensure_empty_row(table_name):
                cur.execute(f"SELECT COUNT(*) FROM {table_name} WHERE [ロットID] = ?", [lot_id])
                cnt_row = cur.fetchone()
                exists = bool(cnt_row and cnt_row[0] and cnt_row[0] > 0)
                if not exists:
                    cur.execute(f"INSERT INTO {table_name} ([ロットID]) VALUES (?)", [lot_id])

            _ensure_empty_row("t_原価見積材料")
            _ensure_empty_row("t_原価見積加工管理")
            _ensure_empty_row("t_原価見積送料")
            _ensure_empty_row("t_原価見積計算チャージ")

        conn.commit()
        return {"ok": True, "estimate_id": estimate_id, "lot_id": lot_id}
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def api_est_calc_delete_estimate_lot(payload=None):
    """見積り削除: 指定ロットIDのデータを関連テーブルから削除し、当該原価見積りIDの初期費用も削除"""
    data = payload or {}
    estimate_id = (data.get("estimate_id") or "").strip()
    lot_id = (data.get("lot_id") or "").strip()

    if not lot_id:
        return {"error": "lot_id が必要です"}

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()

        # 従属テーブル→親テーブルの順に削除（制約がある場合に安全）
        tables = [
            "t_原価見積計算チャージ",
            "t_原価見積送料",
            "t_原価見積加工管理",
            "t_原価見積真鍮",
            "t_原価見積材料",
            "t_原価見積情報",
        ]
        for table_name in tables:
            cur.execute(f"DELETE FROM {table_name} WHERE ロットID = ?", [lot_id])

        if estimate_id:
            cur.execute(
                "DELETE FROM t_原価見積初期費用 WHERE 原価見積りID = ?;",
                [estimate_id],
            )

        conn.commit()
        return {"ok": True, "estimate_id": estimate_id, "lot_id": lot_id}
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def api_est_calc_pre_export_save(payload=None):
    """Excel 出力前の保存処理"""
    data = payload or {}
    if "br_8" in data:
        data["br_8"] = normalize_cutting_recovery_rate_for_db(data.get("br_8"))
    if "mat_17" in data:
        data["mat_17"] = normalize_yield_rate_for_db(data.get("mat_17"))
    estimate_id = (data.get("estimate_id") or "").strip()
    lot_id = (data.get("lot_id") or "").strip()
    if not estimate_id or not lot_id:
        return {"error": "estimate_id と lot_id が必要です"}

    nz_zero_keys = {
        "est_7", "est_8", "est_9", "est_10", "est_11", "est_12",
        "mat_2", "mat_3", "mat_5", "mat_6", "mat_7", "mat_8", "mat_9", "mat_10",
        "mat_11", "mat_12", "mat_13", "mat_14", "mat_15", "mat_16", "mat_17", "mat_18", "mat_19",
        "br_1", "br_2", "br_3", "br_4", "br_5", "br_6", "br_7", "br_8", "br_9", "br_10",
        "proc_1", "proc_2", "proc_3", "proc_4", "proc_5", "proc_6", "proc_7", "proc_8",
        "proc_9", "proc_10", "proc_11", "proc_12", "proc_13", "proc_14", "proc_15", "proc_16", "proc_17",
        "proc_18", "proc_19", "proc_20", "proc_21", "proc_22", "proc_23", "proc_24", "proc_25", "proc_26", "proc_27",
        "proc_28", "proc_29", "proc_30", "proc_32", "proc_33", "proc_34", "proc_35", "proc_36", "proc_37", "proc_38", "proc_39",
        "kensa_1", "kensa_2", "kensa_3", "kensa_4", "kensa_5", "kensa_6", "kensa_7", "kensa_8", "kensa_9",
        "kensa_10", "kensa_11", "kensa_12", "kensa_13", "kensa_14", "kensa_15",
        "kensa_17", "kensa_18", "kensa_19", "kensa_21", "kensa_22", "kensa_24", "kensa_25", "kensa_27", "kensa_28", "kensa_30", "kensa_31",
        "soryo_1", "soryo_2", "soryo_3", "soryo_6", "soryo_7",
        "soryo_9", "soryo_10", "soryo_11", "soryo_12", "soryo_13", "soryo_14", "soryo_15", "soryo_16", "soryo_17", "soryo_18", "soryo_19",
        "soryo_36", "soryo_37", "soryo_38", "soryo_39", "soryo_20", "soryo_21", "soryo_22", "soryo_23",
        "soryo_26", "soryo_27", "soryo_28", "soryo_29", "soryo_30", "soryo_31", "soryo_32", "soryo_33", "soryo_34", "soryo_35",
        "f_in_1", "f_in_2", "f_in_3", "f_out_1", "f_out_2", "f_out_3", "f_out_4", "f_out_5", "f_out_6",
        "f_in_b1", "f_in_b2", "f_in_b3", "f_in_b4", "f_ch_1", "f_ch_2", "f_ch_3", "f_ch_4", "f_ch_5", "f_ch_6", "f_ch_7", "f_ch_8", "f_ch_9", "f_ch_10",
    }

    def _val(key):
        v = data.get(key)
        if key in nz_zero_keys:
            if v is None:
                return "0"
            s = str(v).strip()
            if s == "":
                return "0"
            return s
        if v is None:
            return ""
        return str(v)

    def _run_upsert(cur, table_name, pairs):
        cur.execute(f"SELECT COUNT(*) FROM {table_name} WHERE [ロットID] = ?", [lot_id])
        cnt_row = cur.fetchone()
        exists = bool(cnt_row and cnt_row[0] and cnt_row[0] > 0)

        # if exists:
        #     set_clause = ", ".join([f"[{col}] = ?" for col, _ in pairs])
        #     sql = f"UPDATE {table_name} SET {set_clause} WHERE [ロットID] = ?"
        #     params = [_val(k) for _, k in pairs] + [lot_id]
        #     cur.execute(sql, params)

        # 検証用クエリ
        if exists:
            for col, key in pairs:
                sql = f"UPDATE {table_name} SET [{col}] = ? WHERE [ロットID] = ?"
                params = [_val(key), lot_id]
                try:
                    cur.execute(sql, params)
                except Exception as e:
                    print(f"エラー: {e}")
                    print(f"SQL: {sql}")
                    print(f"Params: {params}")
                    print(f"Table: {table_name}")
                    print(f"Key: {key}")
                    print(f"Value: {_val(key)}")
                    print(f"Lot ID: {lot_id}")
        
        else:
            cols = "[ロットID], " + ", ".join([f"[{col}]" for col, _ in pairs])
            placeholders = ", ".join(["?"] * (len(pairs) + 1))
            sql = f"INSERT INTO {table_name} ({cols} VALUES ({placeholders})"
            params = [lot_id] + [_val(k) for _, k in pairs]
            cur.execute(sql, params)

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()

        # t_原価見積情報
        _run_upsert(cur, "t_原価見積情報", [
            ("ロット数", "est_7"),
            ("CT", "est_8"),
            ("日産数", "est_9"),
            ("日産数入力", "est_10"),
            ("日数補正", "est_12"),
            ("日数", "est_11"),
        ])

        # t_原価見積材料
        _run_upsert(cur, "t_原価見積材料", [
            ("材質径", "mat_1"),
            ("鋼種", "mat_2"),
            ("単重", "mat_3"),
            ("形状", "mat_4"),
            ("径", "mat_5"),
            ("長さ", "mat_6"),
            ("全長", "mat_7"),
            ("突切り", "mat_8"),
            ("取り数", "mat_9"),
            ("取り数入力", "mat_10"),
            ("材料費合計", "mat_11"),
            ("比重", "mat_12"),
            ("一本重", "mat_13"),
            ("一個重", "mat_14"),
            ("材料単価", "mat_15"),
            ("材料費", "mat_16"),
            ("歩留り", "mat_17"),
            ("歩留り金額", "mat_18"),
            ("材料費入力", "mat_19"),
        ])

        # t_原価見積真鍮（有効化チェック時のみ）
        if bool(data.get("br_cb_1")):
            _run_upsert(cur, "t_原価見積真鍮", [
                ("RM区分", "shinchuu_r1"),
                ("重量計算区分", "shinchuu_r2"),
                ("素材単価", "br_1"),
                ("N社価格", "br_2"),
                ("建値", "br_3"),
                ("増値", "br_4"),
                ("真鍮単重", "br_5"),
                ("スクラップ重", "br_6"),
                ("スクラップベース", "br_7"),
                ("切粉回収率", "br_8"),
                ("スクラップ単価", "br_9"),
                ("真鍮材料費", "br_10"),
            ])
        else:
            cur.execute(
                "DELETE FROM t_原価見積真鍮 WHERE [ロットID] = ?",
                [lot_id],
            )

        # t_原価見積加工管理
        _run_upsert(cur, "t_原価見積加工管理", [
            ("刃工具価格", "proc_1"),
            ("材質", "proc_2"),
            ("交換サイクル", "proc_3"),
            ("刃工具原価", "proc_4"),
            ("切削機械", "proc_5"),
            ("切削機械サイクル", "proc_6"),
            ("切削機械賃率", "proc_7"),
            ("切削機械原価", "proc_8"),
            ("ガイドブッシュ切替", "proc_cb_1"),
            ("セット難易度", "proc_9"),
            ("セット時間H", "proc_10"),
            ("セット時間S", "proc_11"),
            ("セット金額", "proc_12"),
            ("セット原価", "proc_13"),
            ("バレル切替", "proc_cb_2"),
            ("バレル", "proc_14"),
            ("バレルサイクル", "proc_15"),
            ("バレル賃率", "proc_16"),
            ("バレル原価", "proc_17"),
            ("ブラスト切替", "proc_cb_3"),
            ("ブラストサイクル", "proc_18"),
            ("ブラスト賃率", "proc_19"),
            ("ブラスト原価", "proc_20"),
            ("圧入切替", "proc_cb_4"),
            ("圧入サイクル", "proc_21"),
            ("圧入賃率", "proc_22"),
            ("圧入原価", "proc_23"),
            ("洗浄", "proc_24"),
            ("洗浄サイクル", "proc_25"),
            ("洗浄賃率", "proc_26"),
            ("洗浄原価", "proc_27"),
            ("処理前検査切替", "proc_cb_5"),
            ("処理前検査サイクル", "proc_28"),
            ("処理前検査賃率", "proc_29"),
            ("処理前検査原価", "proc_30"),
            ("表面処理切替", "proc_cb_6"),
            ("表面処理名ID", "proc_32"),
            ("表面処理原価", "proc_33"),
            ("表面処理名ID2", "proc_34"),
            ("表面処理原価2", "proc_35"),
            ("計量梱包", "proc_36"),
            ("計量梱包サイクル", "proc_37"),
            ("計量梱包賃率", "proc_38"),
            ("計量梱包原価", "proc_39"),
            ("自動外観検査切替", "kensa_cb_1"),
            ("自動外観検査サイクル", "kensa_1"),
            ("自動外観検査賃率", "kensa_2"),
            ("自動外観検査原価", "kensa_3"),
            ("数値切替", "kensa_cb_2"),
            ("数値サイクル", "kensa_4"),
            ("数値賃率", "kensa_5"),
            ("数値原価", "kensa_6"),
            ("目視切替", "kensa_cb_3"),
            ("目視サイクル", "kensa_7"),
            ("目視賃率", "kensa_8"),
            ("目視原価", "kensa_9"),
            ("顕微鏡切替", "kensa_cb_4"),
            ("顕微鏡サイクル", "kensa_10"),
            ("顕微鏡賃率", "kensa_11"),
            ("顕微鏡原価", "kensa_12"),
            ("マイクロゲージ切替", "kensa_cb_5"),
            ("マイクロゲージサイクル", "kensa_13"),
            ("マイクロゲージ賃率", "kensa_14"),
            ("マイクロゲージ原価", "kensa_15"),
            ("その他切替", "kensa_cb_6"),
            ("その他検査名", "kensa_16"),
            ("その他サイクル", "kensa_17"),
            ("その他賃率", "kensa_18"),
            ("その他原価", "kensa_19"),
            ("その他検査名2", "kensa_20"),
            ("その他サイクル2", "kensa_21"),
            ("その他原価2", "kensa_22"),
            ("その他検査名3", "kensa_23"),
            ("その他サイクル3", "kensa_24"),
            ("その他原価3", "kensa_25"),
            ("その他検査名4", "kensa_26"),
            ("その他サイクル4", "kensa_27"),
            ("その他原価4", "kensa_28"),
            ("その他検査名5", "kensa_29"),
            ("その他サイクル5", "kensa_30"),
            ("その他原価5", "kensa_31"),
        ])

        # t_原価見積送料
        _run_upsert(cur, "t_原価見積送料", [
            ("使用箱切替", "est_soryo_box"),
            ("D入数", "soryo_1"),
            ("D納入先", "soryo_2"),
            ("D梱包サイズ", "soryo_3"),
            ("D箱サイズ", "soryo_4"),
            ("D重量", "soryo_5"),
            ("D送料", "soryo_6"),
            ("D運賃単価", "soryo_7"),
            ("D箱規格", "soryo_8"),
            ("D箱価格", "soryo_9"),
            ("D箱入数", "soryo_10"),
            ("D一箱重量", "soryo_11"),
            ("D箱単価", "soryo_12"),
            ("D使用トレー", "soryo_13"),
            ("D収容数", "soryo_14"),
            ("Dトレー価格", "soryo_15"),
            ("D天井フタ数", "soryo_16"),
            ("D必要トレー数", "soryo_17"),
            ("Dトレー合計額", "soryo_18"),
            ("Dトレー単価", "soryo_19"),
            ("納入日数", "soryo_36"),
            ("納入数日", "soryo_37"),
            ("日当重量", "soryo_38"),
            ("日当送料単価", "soryo_39"),
            ("J単価パレット", "soryo_20"),
            ("J搭載箱数パレット", "soryo_21"),
            ("J搭載数パレット", "soryo_22"),
            ("J運賃単価パレット", "soryo_23"),
            ("J箱負担", "soryo_24"),
            ("J箱規格", "soryo_25"),
            ("J箱価格", "soryo_26"),
            ("J箱入数", "soryo_27"),
            ("J箱単価", "soryo_28"),
            ("J使用トレー", "soryo_29"),
            ("J収容数", "soryo_30"),
            ("Jトレー価格", "soryo_31"),
            ("J天井フタ数", "soryo_32"),
            ("J必要トレー数", "soryo_33"),
            ("Jトレー合計額", "soryo_34"),
            ("Jトレー単価", "soryo_35"),
        ])

        # t_原価見積計算チャージ
        _run_upsert(cur, "t_原価見積計算チャージ", [
            ("加工管理合計社内", "f_in_1"),
            ("管理費率社内", "f_in_2"),
            ("管理費社内", "f_in_3"),
            ("表面処理費", "f_out_1"),
            ("管理費率社外", "f_out_2"),
            ("管理費社外", "f_out_3"),
            ("表面処理費2", "f_out_4"),
            ("管理費率社外2", "f_out_5"),
            ("管理費社外2", "f_out_6"),
            ("原価合計", "f_in_b1"),
            ("粗利率", "f_in_b2"),
            ("粗利", "f_in_b3"),
            ("見積単価", "f_in_b4"),
            ("チャージ材料費", "f_ch_1"),
            ("チャージ刃工具費", "f_ch_2"),
            ("チャージ社外管理費", "f_ch_3"),
            ("チャージ表面処理費", "f_ch_4"),
            ("チャージ検査費", "f_ch_5"),
            ("チャージ梱包費", "f_ch_6"),
            ("機械チャージ", "f_ch_7"),
            ("検査チャージ", "f_ch_8"),
            ("梱包チャージ", "f_ch_9"),
            ("チャージ金額", "f_ch_10"),
        ])

        conn.commit()
        return {"ok": True, "estimate_id": estimate_id, "lot_id": lot_id}
    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def api_est_calc_shipping_by_region_size(payload=None):
    """地方ID + 梱包サイズ名から送料を取得する"""
    reg = ((payload or {}).get("reg") or "").strip()
    size_name = ((payload or {}).get("size_name") or "").strip()
    if not reg or not size_name:
        return {"ok": True, "shipping": ""}

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM t_運賃表 WHERE 地方ID = ?", [reg])
        row = cur.fetchone()
        if not row:
            return {"ok": True, "shipping": ""}
        col_names = [c[0] for c in cur.description] if cur.description else []
        rec = dict(zip(col_names, row))
        shipping = rec.get(size_name)
        if shipping is None:
            return {"ok": True, "shipping": ""}
        return {"ok": True, "shipping": str(shipping)}
    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def api_est_calc_initial_cost_row(payload=None):
    """初期費用: 行 ID と原価見積り ID で 1 件取得（フォーム反映用）"""
    estimate_id = ((payload or {}).get("estimate_id") or "").strip()
    row_id = ((payload or {}).get("id") or "").strip()
    if not estimate_id or not row_id:
        return {"error": "estimate_id と id が必要です"}

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT * FROM t_原価見積初期費用 WHERE ID = ? AND 原価見積りID = ?;",
            [row_id, estimate_id],
        )
        ic_row = cur.fetchone()
        if not ic_row:
            return {"error": "該当データが見つかりません"}
        col_names = [c[0] for c in (cur.description or [])]
        out = {}
        for k, v in zip(col_names, ic_row):
            if k in ("数量", "単価", "金額"):
                out[k] = _initial_cost_display_str(k, v)
            elif v is None:
                out[k] = ""
            else:
                out[k] = str(v)
        return {"ok": True, "row": out}
    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def api_est_calc_initial_cost_save(payload=None):
    """初期費用: 新規 INSERT または既存行 UPDATE"""
    data = payload or {}
    estimate_id = (data.get("estimate_id") or "").strip()
    row_id = (data.get("id") or "").strip()
    hinmei = (data.get("品名") or "").strip()
    tani = (data.get("単位") or "").strip()

    if not estimate_id:
        return {"error": "原価見積りIDがありません"}
    if not hinmei:
        return {"error": "品名を入力してください"}
    if tani not in ("個", "式", "セット"):
        return {"error": "単位を選択してください"}

    try:
        suryo = _initial_cost_parse_long_for_save(data.get("数量"), "数量")
        tanka = _initial_cost_parse_long_for_save(data.get("単価"), "単価")
        kingaku_raw = data.get("金額")
        kingaku_s = (
            str(kingaku_raw).replace(",", "").strip()
            if kingaku_raw is not None
            else ""
        )
        if kingaku_s == "":
            kingaku = suryo * tanka
        else:
            kingaku = _initial_cost_parse_long_for_save(kingaku_raw, "金額")
    except ValueError as e:
        return {"error": str(e)}

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        if row_id:
            cur.execute(
                "SELECT 1 FROM t_原価見積初期費用 WHERE ID = ? AND 原価見積りID = ?;",
                [row_id, estimate_id],
            )
            if not cur.fetchone():
                conn.rollback()
                return {"error": "該当データが見つかりません"}
            cur.execute(
                "UPDATE t_原価見積初期費用 SET 品名 = ?, 数量 = ?, 単位 = ?, 単価 = ?, 金額 = ? "
                "WHERE ID = ? AND 原価見積りID = ?;",
                [hinmei, suryo, tani, tanka, kingaku, row_id, estimate_id],
            )
        else:
            cur.execute(
                "INSERT INTO t_原価見積初期費用 (原価見積りID, 品名, 数量, 単位, 単価, 金額) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                [estimate_id, hinmei, suryo, tani, tanka, kingaku],
            )
        conn.commit()
        rows = _initial_cost_rows_for_api(cur, estimate_id)
        return {"ok": True, "rows": rows}
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def api_est_calc_initial_cost_delete(payload=None):
    """初期費用: 1 件削除"""
    data = payload or {}
    estimate_id = (data.get("estimate_id") or "").strip()
    row_id = (data.get("id") or "").strip()
    if not estimate_id or not row_id:
        return {"error": "estimate_id と id が必要です"}

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT 1 FROM t_原価見積初期費用 WHERE ID = ? AND 原価見積りID = ?;",
            [row_id, estimate_id],
        )
        if not cur.fetchone():
            conn.rollback()
            return {"error": "該当データが見つかりません"}
        cur.execute(
            "DELETE FROM t_原価見積初期費用 WHERE ID = ? AND 原価見積りID = ?;",
            [row_id, estimate_id],
        )
        conn.commit()
        rows = _initial_cost_rows_for_api(cur, estimate_id)
        return {"ok": True, "rows": rows}
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def api_est_calc_export_xlsx(payload=None):
    """テンプレートExcelへ値を反映し、xlsx を返却する"""
    data = payload or {}

    def _s(key):
        return str(data.get(key) or "").strip()

    def _is_checked(v):
        s = str(v or "").strip()
        return s in ("-1", "1", "Y", "True", "true")

    def _rate_float_for_formula(key):
        t = _s(key).replace(",", "").replace("％", "").replace("%", "")
        if not t:
            return 0.0
        try:
            return float(t)
        except ValueError:
            return 0.0

    def _excel_num(val):
        """E/F/G 用: 文字列を数値に変換。空・解釈不能は None（セルは空）"""
        if val is None:
            return None
        t = str(val).strip().replace(",", "").replace("，", "").replace(" ", "")
        if not t:
            return None
        try:
            n = float(t)
            if abs(n - round(n)) < 1e-9:
                return int(round(n))
            return n
        except ValueError:
            return None

    def _set_efg(ws, col, row, val):
        n = _excel_num(val)
        ws[f"{col}{row}"] = n

    est_1 = _s("est_1")
    est_2 = _s("est_2")
    est_3 = _s("est_3")
    est_4 = _s("est_4")
    est_5 = _s("est_5")
    est_6 = _s("est_6")
    est_7 = _s("est_7")
    est_11 = _s("est_11")
    mat_1 = _s("mat_1")
    mat_2_name = _s("mat_2_name")
    mat_3 = _s("mat_3")
    mat_15 = _s("mat_15")
    mat_19 = _s("mat_19")
    f_ch_8 = _s("f_ch_8")
    f_ch_9 = _s("f_ch_9")
    proc_5_name = _s("proc_5_name")
    proc_14_name = _s("proc_14_name")
    proc_24_name = _s("proc_24_name")
    proc_36_name = _s("proc_36_name")
    proc_32_name = _s("proc_32_name")
    proc_34_name = _s("proc_34_name")

    customer_part = _sanitize_filename_part(est_3, 8)
    estimate_part = _sanitize_filename_part(est_5)
    lot_part = _sanitize_filename_part(est_6)
    download_name = f"原価見積書_{customer_part}_{estimate_part}_{lot_part}.xlsx"

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl がインストールされていません。pip install openpyxl を実行してください。"}

    if not os.path.exists(EXCEL_TEMPLATE_PATH):
        return {"error": f"テンプレートが見つかりません: {EXCEL_TEMPLATE_PATH}"}

    try:
        wb = load_workbook(EXCEL_TEMPLATE_PATH)
        if "原価見積りシート" not in wb.sheetnames:
            return {"error": "テンプレートに「原価見積りシート」が存在しません。"}
        ws = wb["原価見積りシート"]

        ws["B2"] = est_4
        ws["C5"] = est_5
        ws["C6"] = est_3
        ws["C7"] = est_1
        ws["C8"] = est_2
        ws["C9"] = mat_2_name
        ws["C10"] = mat_1
        ws["C11"] = est_7
        ws["C12"] = mat_3
        _set_efg(ws, "G", 10, f_ch_8)
        _set_efg(ws, "G", 11, f_ch_9)
        _set_efg(ws, "G", 12, est_11)
        ws["D15"] = f"{mat_2_name}     {mat_1}"
        _set_efg(ws, "G", 15, mat_19)
        _f15 = ws["F15"]
        _n15 = _excel_num(mat_15)
        if _n15 is None:
            _f15.value = None
            _f15.number_format = "General"
        else:
            _f15.value = int(round(_n15)) if isinstance(_n15, float) else int(_n15)
            _f15.number_format = "#,##0"

        # D16:G23 可変ブロック
        # D:項目名 / E:サイクル / F:賃率 / G:原価（E〜G は数値、空は None）
        for r in range(16, 24):
            ws[f"D{r}"] = ""
            ws[f"E{r}"] = None
            ws[f"F{r}"] = None
            ws[f"G{r}"] = None

        guide_on = _is_checked(data.get("proc_cb_1"))
        rows = []

        # 切削
        rows.append({
            "label": f"切削({proc_5_name}",
            "cycle": _s("proc_6"),
            "rate": _s("proc_7"),
            "cost": _s("proc_8"),
            "cost_col": "切削機械原価",
        })
        # 刃工具費（賃率列Fに刃工具価格 proc-1、原価列Gに刃工具原価 proc-4）
        rows.append({
            "label": "刃工具費",
            "cycle": "",
            "rate": _s("proc_1"),
            "cost": _s("proc_4"),
            "cost_col": "刃工具原価",
        })
        # セット費（常に表示）
        rows.append({
            "label": f"セット費(ガイド{'有' if guide_on else '無'}, {_s('proc_10')}H, 数量:{est_7}, 費用:{_s('proc_12')})",
            "cycle": "",
            "rate": "",
            "cost": _s("proc_13"),
            "cost_col": "セット原価",
        })
        # バレル（チェックON時）
        if _is_checked(data.get("proc_cb_2")):
            rows.append({
                "label": f"バレル({proc_14_name})",
                "cycle": _s("proc_15"),
                "rate": _s("proc_16"),
                "cost": _s("proc_17"),
                "cost_col": "バレル原価",
            })
        # ブラスト（チェックON時）
        if _is_checked(data.get("proc_cb_3")):
            rows.append({
                "label": "ブラスト",
                "cycle": _s("proc_18"),
                "rate": _s("proc_19"),
                "cost": _s("proc_20"),
                "cost_col": "ブラスト原価",
            })
        # 圧入（チェックON時）
        if _is_checked(data.get("proc_cb_4")):
            rows.append({
                "label": "圧入",
                "cycle": _s("proc_21"),
                "rate": _s("proc_22"),
                "cost": _s("proc_23"),
                "cost_col": "圧入原価",
            })
        # 洗浄
        rows.append({
            "label": f"洗浄({proc_24_name})",
            "cycle": _s("proc_25"),
            "rate": _s("proc_26"),
            "cost": _s("proc_27"),
            "cost_col": "洗浄原価",
        })
        # 計量・梱包
        rows.append({
            "label": f"計量・梱包({proc_36_name})",
            "cycle": _s("proc_37"),
            "rate": _s("proc_38"),
            "cost": _s("proc_39"),
            "cost_col": "計量梱包原価",
        })

        for i, row in enumerate(rows[:8]):
            r = 16 + i
            ws[f"D{r}"] = row["label"]
            _set_efg(ws, "E", r, row["cycle"])
            _set_efg(ws, "F", r, row["rate"])
            _set_efg(ws, "G", r, row["cost"])

        # D24:G33 検査ブロック（最大10行。D34は合計行のため使用しない）
        # D:項目名（その他は入力検査名）/ E:サイクル / F:賃率（その他共通は kensa_18）/ G:原価
        ins_rows = []
        if _is_checked(data.get("proc_cb_5")):
            ins_rows.append({
                "label": "処理前検査",
                "cycle": _s("proc_28"),
                "rate": _s("proc_29"),
                "cost": _s("proc_30"),
                "cost_col": "処理前検査原価",
            })
        if _is_checked(data.get("kensa_cb_1")):
            ins_rows.append({
                "label": "自動外観検査",
                "cycle": _s("kensa_1"),
                "rate": _s("kensa_2"),
                "cost": _s("kensa_3"),
                "cost_col": "自動外観検査原価",
            })
        if _is_checked(data.get("kensa_cb_2")):
            ins_rows.append({
                "label": "数値",
                "cycle": _s("kensa_4"),
                "rate": _s("kensa_5"),
                "cost": _s("kensa_6"),
                "cost_col": "数値原価",
            })
        if _is_checked(data.get("kensa_cb_3")):
            ins_rows.append({
                "label": "目視",
                "cycle": _s("kensa_7"),
                "rate": _s("kensa_8"),
                "cost": _s("kensa_9"),
                "cost_col": "目視原価",
            })
        if _is_checked(data.get("kensa_cb_4")):
            ins_rows.append({
                "label": "顕微鏡",
                "cycle": _s("kensa_10"),
                "rate": _s("kensa_11"),
                "cost": _s("kensa_12"),
                "cost_col": "顕微鏡原価",
            })
        if _is_checked(data.get("kensa_cb_5")):
            ins_rows.append({
                "label": "マイクロゲージ",
                "cycle": _s("kensa_13"),
                "rate": _s("kensa_14"),
                "cost": _s("kensa_15"),
                "cost_col": "マイクロゲージ原価",
            })
        kensa_18_shared = _s("kensa_18")
        if _is_checked(data.get("kensa_cb_6")):
            for name_key, cyc_key, cost_key in (
                ("kensa_16", "kensa_17", "kensa_19"),
                ("kensa_20", "kensa_21", "kensa_22"),
                ("kensa_23", "kensa_24", "kensa_25"),
                ("kensa_26", "kensa_27", "kensa_28"),
                ("kensa_29", "kensa_30", "kensa_31"),
            ):
                nm = _s(name_key)
                cy = _s(cyc_key)
                co = _s(cost_key)
                if nm or cy or co:
                    _other_cost_col_map = {
                        "kensa_19": "その他原価",
                        "kensa_22": "その他原価2",
                        "kensa_25": "その他原価3",
                        "kensa_28": "その他原価4",
                        "kensa_31": "その他原価5",
                    }
                    ins_rows.append({
                        "label": nm,
                        "cycle": cy,
                        "rate": kensa_18_shared,
                        "cost": co,
                        "cost_col": _other_cost_col_map.get(cost_key, ""),
                    })

        # D24:G33 のみクリア。D34/G34 はテンプレの「加工管理合計(社内)」と =SUM(G16:G33) のため変更しない
        for r in range(24, 34):
            ws[f"D{r}"] = ""
            ws[f"E{r}"] = None
            ws[f"F{r}"] = None
            ws[f"G{r}"] = None
        ins_cap = 10
        for i, row in enumerate(ins_rows[:ins_cap]):
            r = 24 + i
            ws[f"D{r}"] = row["label"]
            _set_efg(ws, "E", r, row["cycle"])
            _set_efg(ws, "F", r, row["rate"])
            _set_efg(ws, "G", r, row["cost"])

        # D35:G35 社内管理費（G34＝加工管理合計を参照）
        for r in range(35, 40):
            ws[f"D{r}"] = ""
            ws[f"E{r}"] = None
            ws[f"F{r}"] = None
            ws[f"G{r}"] = None
        f_in_2_disp = _s("f_in_2")
        ws["D35"] = f"管理費(社内,{f_in_2_disp}%)"
        r_in = _rate_float_for_formula("f_in_2")
        ws["G35"] = f"=ROUNDUP(G34*{r_in}/100,2)"

        # D36:G39 社外（表面処理＋管理費）※表面処理名未選択は出力しない、1未選択で2のみのときは上詰め
        surf_blocks = []
        if _s("proc_32"):
            surf_blocks.append(
                {
                    "name": proc_32_name or _s("proc_32"),
                    "cost": _s("proc_33"),
                    "rate_key": "f_out_2",
                }
            )
        if _s("proc_34"):
            surf_blocks.append(
                {
                    "name": proc_34_name or _s("proc_34"),
                    "cost": _s("proc_35"),
                    "rate_key": "f_out_5",
                }
            )
        row_ext = 36
        for idx, blk in enumerate(surf_blocks, start=1):
            rate_disp = _s(blk["rate_key"])
            r_out = _rate_float_for_formula(blk["rate_key"])
            ws[f"D{row_ext}"] = f"表面処理({blk['name']})"
            _set_efg(ws, "G", row_ext, blk["cost"])
            ws[f"D{row_ext + 1}"] = f"管理費({rate_disp}%)"
            # 式を置くセルのみ文字列の数式のまま（他の E/F/G は数値）
            ws[f"G{row_ext + 1}"] = f"=ROUNDUP(G{row_ext}*{r_out}/100,2)"
            row_ext += 2

        # G41 粗利率: 画面上の f-in-b2(%) を 100 で割った値（セル書式が%ならそのまま表示）
        if _s("f_in_b2"):
            ws["G41"] = _rate_float_for_formula("f_in_b2") / 100
        else:
            ws["G41"] = None

        # A44:G46 / J29:M29 送料・梱包（est_soryo_box: 1=段ボール 2=樹脂箱）
        for r in (44, 45, 46):
            for col in "ABCDEFG":
                ws[f"{col}{r}"] = None
        for col in "JKLM":
            ws[f"{col}29"] = None
        s8n = _s("soryo_8_name")
        s13n = _s("soryo_13_name")
        s29n = _s("soryo_29_name")
        box_mode = str(data.get("est_soryo_box") or "").strip()
        if box_mode == "1":
            ws["D44"] = f"運賃単価(入数：{_s('soryo_10')})"
            _set_efg(ws, "G", 44, _s("soryo_7"))
            ws["A45"] = "梱包種類：段ボール"
            ws["D45"] = f"箱単価({s8n or _s('soryo_8')})"
            _set_efg(ws, "G", 45, _s("soryo_12"))
            ws["D46"] = f"トレー単価({s13n or _s('soryo_13')})"
            _set_efg(ws, "G", 46, _s("soryo_19"))
        elif box_mode == "2":
            ws["D44"] = f"運賃単価(入数：{_s('soryo_27')})"
            _set_efg(ws, "G", 44, _s("soryo_23"))
            ws["A45"] = "梱包種類：パレット"
            if _s("soryo_24") == "自達":
                ws["D45"] = f"箱単価({_s('soryo_25')})"
            else:
                ws["D45"] = "箱単価(箱支給)"
            _set_efg(ws, "G", 45, _s("soryo_28"))
            ws["J29"] = f"トレー単価({s29n or _s('soryo_29')})"
            _set_efg(ws, "M", 29, _s("soryo_35"))

        def _txt(v):
            if v is None:
                return ""
            return str(v).strip()

        def _parse_float_loose(v):
            if v is None:
                return None
            t = str(v).strip().replace(",", "").replace("，", "").replace(" ", "")
            if not t:
                return None
            try:
                return float(t)
            except ValueError:
                return None

        def _shipping_sum_from_row_dict(row_dict):
            """履歴1行: 送料単価 = 運賃単価 + 箱単価 + トレー単価（欠損は0扱い）。使用箱切替で列を切替。"""
            box = str(row_dict.get("使用箱切替") or "").strip()

            def _z(key):
                x = _parse_float_loose(row_dict.get(key))
                return 0.0 if x is None else x

            if box == "1":
                return _z("D運賃単価") + _z("D箱単価") + _z("Dトレー単価")
            if box == "2":
                return _z("J運賃単価パレット") + _z("J箱単価") + _z("Jトレー単価")
            return None

        def _db_text(row_dict, col_name):
            if not col_name:
                return ""
            if col_name == "__CHARGE_TOTAL__":
                def _to_float(_v):
                    t = str(_v or "").strip().replace(",", "")
                    if not t:
                        return None
                    try:
                        return float(t)
                    except ValueError:
                        return None
                _m = _to_float(row_dict.get("機械チャージ"))
                _k = _to_float(row_dict.get("検査チャージ"))
                _p = _to_float(row_dict.get("梱包チャージ"))
                if _m is not None and _k is not None and _p is not None:
                    _sum = _m + _k + _p
                    if abs(_sum - round(_sum)) < 1e-9:
                        return str(int(round(_sum)))
                    return f"{_sum:.2f}".rstrip("0").rstrip(".")
                return _txt(row_dict.get("チャージ金額"))
            if col_name == "__SHIPPING_UNIT__":
                _ship = _shipping_sum_from_row_dict(row_dict)
                if _ship is None:
                    return ""
                return str(_ship)
            if col_name == "__TOTAL_UNIT__":
                _unit = _parse_float_loose(row_dict.get("見積単価"))
                _ship = _shipping_sum_from_row_dict(row_dict)
                if _unit is None or _ship is None:
                    return ""
                _sum = _unit + _ship
                return str(_sum)
            return _txt(row_dict.get(col_name))

        def _sum3_text(a, b, c, fallback=""):
            def _to_float(_v):
                t = str(_v or "").strip().replace(",", "")
                if not t:
                    return None
                try:
                    return float(t)
                except ValueError:
                    return None
            _a = _to_float(a)
            _b = _to_float(b)
            _c = _to_float(c)
            if _a is None or _b is None or _c is None:
                return _txt(fallback)
            _sum = _a + _b + _c
            if abs(_sum - round(_sum)) < 1e-9:
                return str(int(round(_sum)))
            return f"{_sum:.2f}".rstrip("0").rstrip(".")

        def _set_cmp_cell(cell, raw_val, number_format):
            if number_format == "@":
                _v = _txt(raw_val)
                cell.value = None if _v == "" else _v
                cell.number_format = "@"
                return
            _n = _excel_num(raw_val)
            cell.value = _n
            cell.number_format = number_format

        cmp_top_rows = [
            {"row": 1, "d_value": est_6, "db_col": "ロットID", "fmt": "@"},
            {"row": 3, "d_value": est_7, "db_col": "ロット数", "fmt": "#,##0"},
            {"row": 4, "d_value": _s("f_ch_1"), "db_col": "チャージ材料費", "fmt": "#,##0.00"},
            {"row": 5, "d_value": _s("f_ch_2"), "db_col": "チャージ刃工具費", "fmt": "#,##0.00"},
            {"row": 6, "d_value": _s("f_ch_4"), "db_col": "チャージ表面処理費", "fmt": "#,##0.00"},
            {"row": 7, "d_value": _s("f_ch_3"), "db_col": "チャージ社外管理費", "fmt": "#,##0.00"},
            {"row": 8, "d_value": _s("f_ch_5"), "db_col": "チャージ検査費", "fmt": "#,##0.00"},
            {"row": 9, "d_value": _s("f_ch_6"), "db_col": "チャージ梱包費", "fmt": "#,##0.00"},
            {"row": 11, "d_value": _s("f_ch_7"), "db_col": "機械チャージ", "fmt": "#,##0.00"},
            {"row": 12, "d_value": _s("f_ch_8"), "db_col": "検査チャージ", "fmt": "#,##0.00"},
            {"row": 13, "d_value": _s("f_ch_9"), "db_col": "梱包チャージ", "fmt": "#,##0.00"},
            {"row": 14, "d_value": est_11, "db_col": "日数", "fmt": "#,##0.00"},
            {"row": 15, "d_value": _sum3_text(_s("f_ch_7"), _s("f_ch_8"), _s("f_ch_9"), _s("f_ch_10")), "db_col": "__CHARGE_TOTAL__", "fmt": "#,##0.00"},
        ]

        # 比較表シート C17〜: 項目列 + D列(画面値) + E列以降(履歴クエリ値)
        cmp_rows = []
        cmp_rows.append({"label": f"{mat_2_name}     {mat_1}", "d_value": _s("mat_19"), "db_col": "材料費入力"})
        cmp_rows.append({"label": "材料単価", "d_value": mat_15, "db_col": "材料単価", "fmt": "#,##0"})
        for row in rows[:8]:
            cmp_rows.append({"label": row["label"], "d_value": row["cost"], "db_col": row.get("cost_col", "")})
        if ins_rows:
            cmp_rows.append({"label": "", "d_value": "", "db_col": ""})
        for row in ins_rows[:ins_cap]:
            cmp_rows.append({"label": row["label"], "d_value": row["cost"], "db_col": row.get("cost_col", "")})
        cmp_rows.append({"label": "", "d_value": "", "db_col": ""})
        cmp_rows.append({"label": "加工管理合計(社内)", "d_value": _s("f_in_1"), "db_col": "加工管理合計社内"})
        cmp_rows.append({"label": "管理費率(社内)", "d_value": _s("f_in_2"), "db_col": "管理費率社内"})
        cmp_rows.append({"label": "管理費(社内)", "d_value": _s("f_in_3"), "db_col": "管理費社内"})
        if surf_blocks:
            cmp_rows.append({"label": "", "d_value": "", "db_col": ""})
            for idx, blk in enumerate(surf_blocks, start=1):
                if idx == 1:
                    cmp_rows.append({"label": f"表面処理({blk['name']})", "d_value": _s("proc_33"), "db_col": "表面処理原価"})
                    cmp_rows.append({"label": f"管理費率(社外{idx})", "d_value": _s("f_out_2"), "db_col": "管理費率社外"})
                    cmp_rows.append({"label": f"管理費(社外{idx})", "d_value": _s("f_out_3"), "db_col": "管理費社外"})
                elif idx == 2:
                    cmp_rows.append({"label": f"表面処理({blk['name']})", "d_value": _s("proc_35"), "db_col": "表面処理原価2"})
                    cmp_rows.append({"label": f"管理費率(社外{idx})", "d_value": _s("f_out_5"), "db_col": "管理費率社外2"})
                    cmp_rows.append({"label": f"管理費(社外{idx})", "d_value": _s("f_out_6"), "db_col": "管理費社外2"})
        cmp_rows.append({"label": "", "d_value": "", "db_col": ""})
        cmp_rows.append({"label": "原価合計", "d_value": _s("f_in_b1"), "db_col": "原価合計"})
        cmp_rows.append({"label": "粗利率", "d_value": _s("f_in_b2"), "db_col": "粗利率"})
        cmp_rows.append({"label": "粗利", "d_value": _s("f_in_b3"), "db_col": "粗利"})
        cmp_rows.append({"label": "見積単価", "d_value": _s("f_in_b4"), "db_col": "見積単価"})
        cmp_rows.append({"label": "", "d_value": "", "db_col": ""})
        if box_mode == "1":
            cmp_rows.append({"label": "運賃単価", "d_value": _s("soryo_7"), "db_col": "D運賃単価"})
            cmp_rows.append({"label": "箱単価", "d_value": _s("soryo_12"), "db_col": "D箱単価"})
            _tray_use = bool((s13n or "").strip() or (_s("soryo_13") or "").strip())
            if _tray_use:
                cmp_rows.append({"label": "トレー単価", "d_value": _s("soryo_19"), "db_col": "Dトレー単価"})
            cmp_rows.append({"label": "送料単価", "d_value": _s("f_in_b5"), "db_col": "__SHIPPING_UNIT__"})
        elif box_mode == "2":
            cmp_rows.append({"label": "運賃単価", "d_value": _s("soryo_23"), "db_col": "J運賃単価パレット"})
            cmp_rows.append({"label": "箱単価", "d_value": _s("soryo_28"), "db_col": "J箱単価"})
            _tray_use2 = bool((s29n or "").strip() or (_s("soryo_29") or "").strip())
            if _tray_use2:
                cmp_rows.append({"label": "トレー単価", "d_value": _s("soryo_35"), "db_col": "Jトレー単価"})
            cmp_rows.append({"label": "送料単価", "d_value": _s("f_in_b5"), "db_col": "__SHIPPING_UNIT__"})
        cmp_rows.append({"label": "", "d_value": "", "db_col": ""})
        cmp_rows.append({"label": "合計単価", "d_value": _s("f_in_b6"), "db_col": "__TOTAL_UNIT__"})

        # 「比較表」シート: C17〜 項目列 / D列=画面値 / E列以降=履歴ロット値
        if "比較表" in wb.sheetnames:
            ws_cmp = wb["比較表"]
            for _cc in range(4, 121):
                for _rr in (1, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14, 15):
                    _vclr = ws_cmp.cell(row=_rr, column=_cc)
                    _vclr.value = None
                    _vclr.number_format = "General"
            for _rr in range(17, 400):
                _cclr = ws_cmp.cell(row=_rr, column=3)
                _cclr.value = None
                _cclr.number_format = "General"
            for _cc in range(4, 121):
                for _rr in range(17, 400):
                    _vclr = ws_cmp.cell(row=_rr, column=_cc)
                    _vclr.value = None
                    _vclr.number_format = "General"
            for _top in cmp_top_rows:
                _dcel = ws_cmp.cell(row=_top["row"], column=4)
                _set_cmp_cell(_dcel, _top["d_value"], _top["fmt"])
            _r_c = 17
            for _row in cmp_rows:
                _ccel = ws_cmp.cell(row=_r_c, column=3)
                _ccel.value = None if _row["label"] == "" else str(_row["label"])
                _ccel.number_format = "@"
                _dcel = ws_cmp.cell(row=_r_c, column=4)
                _set_cmp_cell(_dcel, _row["d_value"], _row.get("fmt", "#,##0.00"))
                _r_c += 1

            if est_5:
                _cmp_start_col = 5
                _cmp_clear_max_col = 120

                conn_h = None
                try:
                    conn_h = get_connection()
                    cur_h = conn_h.cursor()
                    sql_h = (
                        "SELECT * FROM (((((((t_原価見積履歴 "
                        "LEFT JOIN t_原価見積情報 ON t_原価見積履歴.原価見積りID = t_原価見積情報.原価見積りID) "
                        "LEFT JOIN t_原価見積材料 ON t_原価見積情報.ロットID = t_原価見積材料.ロットID) "
                        "LEFT JOIN t_原価見積真鍮 ON t_原価見積情報.ロットID = t_原価見積真鍮.ロットID) "
                        "LEFT JOIN t_原価見積加工管理 ON t_原価見積情報.ロットID = t_原価見積加工管理.ロットID) "
                        "LEFT JOIN t_原価見積送料 ON t_原価見積情報.ロットID = t_原価見積送料.ロットID) "
                        "LEFT JOIN t_原価見積計算チャージ ON t_原価見積情報.ロットID = t_原価見積計算チャージ.ロットID) "
                        "LEFT JOIN t_営業マスタ ON t_原価見積履歴.営業ID = t_営業マスタ.コード) "
                        "LEFT JOIN t_客先マスタ ON t_原価見積履歴.客先コード = t_客先マスタ.コード "
                        "WHERE t_原価見積履歴.原価見積りID = ? "
                        "AND t_原価見積情報.使用フラグ <> 'Y' "
                        "ORDER BY t_原価見積情報.ロットID ASC;"
                    )
                    cur_h.execute(sql_h, [est_5])
                    _lot_rows = cur_h.fetchall()
                    _lot_cols = [d[0] for d in cur_h.description]
                finally:
                    if conn_h is not None:
                        try:
                            conn_h.close()
                        except Exception:
                            pass

                for _i, _row in enumerate(_lot_rows):
                    _col = _cmp_start_col + _i
                    if _col > _cmp_clear_max_col:
                        break
                    if not _row:
                        continue
                    _row_dict = {_lot_cols[_j]: _row[_j] for _j in range(min(len(_lot_cols), len(_row)))}
                    for _top in cmp_top_rows:
                        _tcel = ws_cmp.cell(row=_top["row"], column=_col)
                        _tv = _db_text(_row_dict, _top["db_col"])
                        _set_cmp_cell(_tcel, _tv, _top["fmt"])
                    _r_v = 17
                    for _cmp in cmp_rows:
                        _vcel = ws_cmp.cell(row=_r_v, column=_col)
                        _v = _db_text(_row_dict, _cmp["db_col"])
                        _set_cmp_cell(_vcel, _v, _cmp.get("fmt", "#,##0.00"))
                        _r_v += 1

        # 原価見積初期費用（1件以上あるときのみ C51 と D〜G 列 51 行目から）
        if est_5:
            ic_rows = []
            conn_ic = None
            try:
                conn_ic = get_connection()
                cur_ic = conn_ic.cursor()
                cur_ic.execute(
                    "SELECT 品名, 数量, 単位, 金額 FROM t_原価見積初期費用 WHERE 原価見積りID = ?",
                    [est_5],
                )
                ic_rows = cur_ic.fetchall()
            finally:
                if conn_ic is not None:
                    try:
                        conn_ic.close()
                    except Exception:
                        pass

            if ic_rows:
                start_r = 51
                for i, ic_row in enumerate(ic_rows):
                    r = start_r + i
                    hinmei = "" if not ic_row or ic_row[0] is None else str(ic_row[0])
                    suryo = ic_row[1] if len(ic_row) > 1 else None
                    tani = "" if len(ic_row) < 3 or ic_row[2] is None else str(ic_row[2])
                    kingaku = ic_row[3] if len(ic_row) > 3 else None
                    ws[f"D{r}"] = hinmei
                    _set_efg(ws, "E", r, suryo)
                    ws[f"F{r}"] = tani
                    _set_efg(ws, "G", r, kingaku)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return {
            "ok": True,
            "_xlsx_bytes": output.getvalue(),
            "_xlsx_name": download_name,
        }
    except Exception as e:
        return {"error": str(e)}


def get_search_page(payload=None):
    """営業担当・客先名のプルダウン用データを取得して search.html を表示"""
    sales_list = []
    customer_list = []
    try:
        conn = get_connection()
        cur = conn.cursor()

        # 営業担当（表示用: 営業担当, 絞り込み用: コード）
        cur.execute(
            "SELECT 営業担当, コード FROM t_営業マスタ WHERE 表示フラグ = 'Y'"
        )
        rows = cur.fetchall()
        sales_list = [
            {"name": row[0], "code": row[1]}
            for row in rows
            if row[0] and row[1]
        ]

        # 客先名（表示用: 客先名, 絞り込み用: コード）
        cur.execute(
            "SELECT 客先名, コード FROM t_客先マスタ WHERE 表示フラグ = 'Y' ORDER BY かな"
        )
        rows = cur.fetchall()
        customer_list = [
            {"name": row[0], "code": row[1]}
            for row in rows
            if row[0] and row[1]
        ]

        conn.close()
    except Exception:
        import traceback
        traceback.print_exc()

    return {
        "ok": True,
        "sales_list": sales_list,
        "customer_list": customer_list,
    }


def _json_safe_rate_master_cell(v):
    """pyodbc セルを JSON 向けに変換"""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _rate_master_row_to_dict(row, col_names):
    return {
        col_names[i]: _json_safe_rate_master_cell(row[i])
        for i in range(len(col_names))
    }


def _compute_rate_sum_from_row_dict(d):
    """労務費賃率～土地の合計を小数第3位（賃率計）"""
    keys = ("労務費賃率", "油等賃率", "電気賃率", "設備費賃率", "建屋", "土地")
    total = Decimal("0")
    for k in keys:
        v = d.get(k)
        if v is None or v == "":
            continue
        try:
            total += Decimal(str(v).replace(",", ""))
        except (InvalidOperation, ValueError, TypeError):
            continue
    try:
        q = total.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        return float(q)
    except (InvalidOperation, ValueError, TypeError):
        return 0.0


def _parse_optional_rate_number(raw):
    """賃率マスタの数値列: 空欄は None"""
    if raw is None:
        return None
    s = str(raw).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        raise ValueError("数値にできません: " + str(raw))


def api_rate_master_list(payload=None):
    """t_賃率マスタ全件（編集画面用）"""
    sql = (
        "SELECT ID, 工程分類, 設備名等, [労務費賃率], [油等賃率], [電気賃率], [設備費賃率], [建屋], [土地] "
        "FROM t_賃率マスタ ORDER BY ID"
    )
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql)
        rows = cur.fetchall()
        col_names = [c[0] for c in cur.description]
        conn.close()
        out_rows = []
        for r in rows:
            d = _rate_master_row_to_dict(r, col_names)
            d["賃率計"] = _compute_rate_sum_from_row_dict(d)
            out_rows.append(d)
        display_columns = list(col_names) + ["賃率計"]
        return {"rows": out_rows, "columns": display_columns}
    except Exception as e:
        return {"error": str(e)}


def _rate_master_payload_from_json(data):
    """リクエスト JSON から更新用タプル (工程分類, 設備名等, 六数値...) とエラー文字列"""
    工程分類 = (data.get("工程分類") or "").strip()
    設備名等 = (data.get("設備名等") or "").strip()
    try:
        労務費賃率 = _parse_optional_rate_number(data.get("労務費賃率"))
        油等賃率 = _parse_optional_rate_number(data.get("油等賃率"))
        電気賃率 = _parse_optional_rate_number(data.get("電気賃率"))
        設備費賃率 = _parse_optional_rate_number(data.get("設備費賃率"))
        建屋 = _parse_optional_rate_number(data.get("建屋"))
        土地 = _parse_optional_rate_number(data.get("土地"))
    except ValueError as ex:
        return None, str(ex)
    params = (
        工程分類 if 工程分類 else None,
        設備名等 if 設備名等 else None,
        労務費賃率,
        油等賃率,
        電気賃率,
        設備費賃率,
        建屋,
        土地,
    )
    return params, None


def api_rate_master_save(payload=None):
    """t_賃率マスタ: ID ありなら UPDATE、なければ INSERT（ID はオートナンバー）"""
    data = payload or {}
    rid = data.get("ID")
    params, err = _rate_master_payload_from_json(data)
    if err:
        return {"error": err}
    if params is None:
        return {"error": "入力を確認してください"}

    is_insert = rid is None or str(rid).strip() == ""

    try:
        conn = get_connection()
        cur = conn.cursor()
        if is_insert:
            insert_sql = (
                "INSERT INTO t_賃率マスタ ([工程分類], [設備名等], [労務費賃率], [油等賃率], [電気賃率], [設備費賃率], [建屋], [土地]) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)"
            )
            cur.execute(insert_sql, params)
            conn.commit()
            cur.execute("SELECT MAX(ID) FROM t_賃率マスタ")
            mx = cur.fetchone()
            new_id = mx[0] if mx else None
            conn.close()
            if new_id is None:
                return {"error": "登録後のIDを取得できませんでした"}
            return {"ok": True, "id": new_id, "inserted": True}
        update_sql = (
            "UPDATE t_賃率マスタ SET "
            "[工程分類]=?, [設備名等]=?, [労務費賃率]=?, [油等賃率]=?, [電気賃率]=?, [設備費賃率]=?, [建屋]=?, [土地]=? "
            "WHERE ID=?"
        )
        try:
            rid_param = int(str(rid).strip())
        except ValueError:
            rid_param = str(rid).strip()
        cur.execute(update_sql, params + (rid_param,))
        if cur.rowcount == 0:
            conn.close()
            return {"error": "該当IDの行がありません"}
        conn.commit()
        conn.close()
        return {"ok": True, "id": rid_param, "inserted": False}
    except Exception as e:
        return {"error": str(e)}


def api_rate_master_delete(payload=None):
    """t_賃率マスタ 1 行削除"""
    data = payload or {}
    rid = data.get("ID")
    if rid is None or str(rid).strip() == "":
        return {"error": "IDが必要です"}
    try:
        rid_param = int(str(rid).strip())
    except ValueError:
        rid_param = str(rid).strip()
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM t_賃率マスタ WHERE ID=?", (rid_param,))
        if cur.rowcount == 0:
            conn.close()
            return {"error": "該当IDの行がありません"}
        conn.commit()
        conn.close()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


def _parse_optional_freight_number(raw):
    """運賃表のサイズ列: 空欄は None、整数のみ（カンマ可）"""
    if raw is None:
        return None
    s = str(raw).strip().replace(",", "")
    if s == "":
        return None
    try:
        d = Decimal(s)
        if d != d.to_integral_value(rounding=ROUND_HALF_UP):
            raise ValueError("小数は入力できません: " + str(raw))
        return int(d)
    except (InvalidOperation, ValueError):
        raise ValueError("整数にできません: " + str(raw))


def _freight_payload_from_json(data):
    """リクエスト JSON から更新/登録用タプルを生成"""
    地方名 = (data.get("地方名") or "").strip()
    try:
        v60 = _parse_optional_freight_number(data.get("60サイズ"))
        v80 = _parse_optional_freight_number(data.get("80サイズ"))
        v100 = _parse_optional_freight_number(data.get("100サイズ"))
        v120 = _parse_optional_freight_number(data.get("120サイズ"))
        v140 = _parse_optional_freight_number(data.get("140サイズ"))
        v160 = _parse_optional_freight_number(data.get("160サイズ"))
        v180 = _parse_optional_freight_number(data.get("180サイズ"))
        v200 = _parse_optional_freight_number(data.get("200サイズ"))
    except ValueError as ex:
        return None, str(ex)
    params = (
        地方名 if 地方名 else None,
        v60, v80, v100, v120, v140, v160, v180, v200,
    )
    return params, None


def api_freight_master_list(payload=None):
    """t_運賃表 一覧"""
    sql = (
        "SELECT [地方ID], [地方名], [60サイズ], [80サイズ], [100サイズ], [120サイズ], [140サイズ], [160サイズ], [180サイズ], [200サイズ] "
        "FROM t_運賃表 ORDER BY [地方ID]"
    )
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql)
        rows = cur.fetchall()
        col_names = [c[0] for c in cur.description]
        conn.close()
        out_rows = [_rate_master_row_to_dict(r, col_names) for r in rows]
        return {"rows": out_rows, "columns": col_names}
    except Exception as e:
        return {"error": str(e)}


def api_freight_master_save(payload=None):
    """t_運賃表: 地方ID 空なら INSERT、ありなら UPDATE"""
    data = payload or {}
    region_id = data.get("地方ID")
    params, err = _freight_payload_from_json(data)
    if err:
        return {"error": err}
    if params is None:
        return {"error": "入力を確認してください"}

    is_insert = region_id is None or str(region_id).strip() == ""

    try:
        conn = get_connection()
        cur = conn.cursor()
        if is_insert:
            insert_sql = (
                "INSERT INTO t_運賃表 ([地方名], [60サイズ], [80サイズ], [100サイズ], [120サイズ], [140サイズ], [160サイズ], [180サイズ], [200サイズ]) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)"
            )
            cur.execute(insert_sql, params)
            conn.commit()
            cur.execute("SELECT MAX([地方ID]) FROM t_運賃表")
            mx = cur.fetchone()
            new_id = mx[0] if mx else None
            conn.close()
            if new_id is None:
                return {"error": "登録後の地方IDを取得できませんでした"}
            return {"ok": True, "id": new_id, "inserted": True}

        update_sql = (
            "UPDATE t_運賃表 SET "
            "[地方名]=?, [60サイズ]=?, [80サイズ]=?, [100サイズ]=?, [120サイズ]=?, [140サイズ]=?, [160サイズ]=?, [180サイズ]=?, [200サイズ]=? "
            "WHERE [地方ID]=?"
        )
        try:
            region_id_param = int(str(region_id).strip())
        except ValueError:
            region_id_param = str(region_id).strip()
        cur.execute(update_sql, params + (region_id_param,))
        if cur.rowcount == 0:
            conn.close()
            return {"error": "該当する地方IDがありません"}
        conn.commit()
        conn.close()
        return {"ok": True, "id": region_id_param, "inserted": False}
    except Exception as e:
        return {"error": str(e)}


def api_freight_master_delete(payload=None):
    """t_運賃表 1 行削除"""
    data = payload or {}
    region_id = data.get("地方ID")
    if region_id is None or str(region_id).strip() == "":
        return {"error": "地方IDが必要です"}
    try:
        region_id_param = int(str(region_id).strip())
    except ValueError:
        region_id_param = str(region_id).strip()
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM t_運賃表 WHERE [地方ID]=?", (region_id_param,))
        if cur.rowcount == 0:
            conn.close()
            return {"error": "該当する地方IDがありません"}
        conn.commit()
        conn.close()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


def _parse_optional_tray_capacity(raw):
    """t_トレー 収容数: 空欄は None、整数のみ（カンマ可）"""
    if raw is None:
        return None
    s = str(raw).strip().replace(",", "")
    if s == "":
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        raise ValueError("収容数を整数にできません: " + str(raw))
    if d != d.to_integral_value(rounding=ROUND_HALF_UP):
        raise ValueError("収容数は整数で入力してください: " + str(raw))
    return int(d)


def _parse_optional_tray_unit_price(raw):
    """t_トレー 単価: 空欄は None、小数第1位まで（カンマ可）"""
    if raw is None:
        return None
    s = str(raw).strip().replace(",", "")
    if s == "":
        return None
    try:
        d = Decimal(s)
        q = d.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
        return float(q)
    except (InvalidOperation, ValueError):
        raise ValueError("単価を数値にできません: " + str(raw))


def _tray_payload_from_json(data):
    """リクエスト JSON から更新/登録用タプルを生成"""
    トレー名 = (data.get("トレー名") or "").strip()
    材質 = (data.get("材質") or "").strip()
    try:
        収容数 = _parse_optional_tray_capacity(data.get("収容数"))
        単価 = _parse_optional_tray_unit_price(data.get("単価"))
    except ValueError as ex:
        return None, str(ex)
    params = (
        トレー名 if トレー名 else None,
        材質 if 材質 else None,
        収容数,
        単価,
    )
    return params, None


def api_tray_master_list(payload=None):
    """t_トレー 一覧"""
    sql = (
        "SELECT ID, [トレー名], [材質], [収容数], [単価] FROM t_トレー ORDER BY ID"
    )
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql)
        rows = cur.fetchall()
        col_names = [c[0] for c in cur.description]
        conn.close()
        out_rows = [_rate_master_row_to_dict(r, col_names) for r in rows]
        return {"rows": out_rows, "columns": col_names}
    except Exception as e:
        return {"error": str(e)}


def api_tray_master_save(payload=None):
    """t_トレー: ID 空なら INSERT、ありなら UPDATE（ID はオートナンバー）"""
    data = payload or {}
    rid = data.get("ID")
    params, err = _tray_payload_from_json(data)
    if err:
        return {"error": err}
    if params is None:
        return {"error": "入力を確認してください"}

    is_insert = rid is None or str(rid).strip() == ""

    try:
        conn = get_connection()
        cur = conn.cursor()
        if is_insert:
            insert_sql = (
                "INSERT INTO t_トレー ([トレー名], [材質], [収容数], [単価]) "
                "VALUES (?, ?, ?, ?)"
            )
            cur.execute(insert_sql, params)
            conn.commit()
            cur.execute("SELECT MAX(ID) FROM t_トレー")
            mx = cur.fetchone()
            new_id = mx[0] if mx else None
            conn.close()
            if new_id is None:
                return {"error": "登録後のIDを取得できませんでした"}
            return {"ok": True, "id": new_id, "inserted": True}

        update_sql = (
            "UPDATE t_トレー SET [トレー名]=?, [材質]=?, [収容数]=?, [単価]=? WHERE ID=?"
        )
        try:
            rid_param = int(str(rid).strip())
        except ValueError:
            rid_param = str(rid).strip()
        cur.execute(update_sql, params + (rid_param,))
        if cur.rowcount == 0:
            conn.close()
            return {"error": "該当IDの行がありません"}
        conn.commit()
        conn.close()
        return {"ok": True, "id": rid_param, "inserted": False}
    except Exception as e:
        return {"error": str(e)}


def api_tray_master_delete(payload=None):
    """t_トレー 1 行削除"""
    data = payload or {}
    rid = data.get("ID")
    if rid is None or str(rid).strip() == "":
        return {"error": "IDが必要です"}
    try:
        rid_param = int(str(rid).strip())
    except ValueError:
        rid_param = str(rid).strip()
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM t_トレー WHERE ID=?", (rid_param,))
        if cur.rowcount == 0:
            conn.close()
            return {"error": "該当IDの行がありません"}
        conn.commit()
        conn.close()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


def _dbox_payload_from_json(data):
    """t_段ボール リクエスト JSON から更新/登録用タプル"""
    規格 = (data.get("規格") or "").strip()
    try:
        単価 = _parse_optional_tray_unit_price(data.get("単価"))
    except ValueError as ex:
        return None, str(ex)
    params = (規格 if 規格 else None, 単価)
    return params, None


def api_dbox_master_list(payload=None):
    """t_段ボール 一覧"""
    sql = "SELECT ID, [規格], [単価] FROM t_段ボール ORDER BY ID"
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql)
        rows = cur.fetchall()
        col_names = [c[0] for c in cur.description]
        conn.close()
        out_rows = [_rate_master_row_to_dict(r, col_names) for r in rows]
        return {"rows": out_rows, "columns": col_names}
    except Exception as e:
        return {"error": str(e)}


def api_dbox_master_save(payload=None):
    """t_段ボール: ID 空なら INSERT、ありなら UPDATE（ID はオートナンバー）"""
    data = payload or {}
    rid = data.get("ID")
    params, err = _dbox_payload_from_json(data)
    if err:
        return {"error": err}
    if params is None:
        return {"error": "入力を確認してください"}

    is_insert = rid is None or str(rid).strip() == ""

    try:
        conn = get_connection()
        cur = conn.cursor()
        if is_insert:
            insert_sql = "INSERT INTO t_段ボール ([規格], [単価]) VALUES (?, ?)"
            cur.execute(insert_sql, params)
            conn.commit()
            cur.execute("SELECT MAX(ID) FROM t_段ボール")
            mx = cur.fetchone()
            new_id = mx[0] if mx else None
            conn.close()
            if new_id is None:
                return {"error": "登録後のIDを取得できませんでした"}
            return {"ok": True, "id": new_id, "inserted": True}

        update_sql = "UPDATE t_段ボール SET [規格]=?, [単価]=? WHERE ID=?"
        try:
            rid_param = int(str(rid).strip())
        except ValueError:
            rid_param = str(rid).strip()
        cur.execute(update_sql, params + (rid_param,))
        if cur.rowcount == 0:
            conn.close()
            return {"error": "該当IDの行がありません"}
        conn.commit()
        conn.close()
        return {"ok": True, "id": rid_param, "inserted": False}
    except Exception as e:
        return {"error": str(e)}


def api_dbox_master_delete(payload=None):
    """t_段ボール 1 行削除"""
    data = payload or {}
    rid = data.get("ID")
    if rid is None or str(rid).strip() == "":
        return {"error": "IDが必要です"}
    try:
        rid_param = int(str(rid).strip())
    except ValueError:
        rid_param = str(rid).strip()
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM t_段ボール WHERE ID=?", (rid_param,))
        if cur.rowcount == 0:
            conn.close()
            return {"error": "該当IDの行がありません"}
        conn.commit()
        conn.close()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


def api_search_conditions(payload=None):
    """search.html からの条件指定検索API"""
    sales_id = ((payload or {}).get('sales_id') or '').strip()
    customer_code = ((payload or {}).get('customer_code') or '').strip()
    part_no = ((payload or {}).get('part_no') or '').strip()
    part_name = ((payload or {}).get('part_name') or '').strip()
    estimate_id = ((payload or {}).get('estimate_id') or '').strip()

    # すべて未指定ならフロント側で弾く想定だが、念のためチェック
    if not any([sales_id, customer_code, part_no, part_name, estimate_id]):
        return {"error": "条件を最低1つ指定してください"}

    base_sql = (
        "SELECT "
        "t_原価見積履歴.原価見積りID, "
        "t_原価見積履歴.管理NO, "
        "t_営業マスタ.営業担当, "
        "t_客先マスタ.客先名, "
        "t_原価見積履歴.品番, "
        "t_原価見積履歴.品名, "
        "t_原価見積履歴.備考 "
        "FROM ((t_原価見積履歴 LEFT JOIN t_営業マスタ ON t_原価見積履歴.営業ID = t_営業マスタ.コード) "
        "LEFT JOIN t_客先マスタ ON t_原価見積履歴.客先コード = t_客先マスタ.コード)"
    )

    conditions = []
    params = []

    if sales_id:
        conditions.append("t_原価見積履歴.営業ID = ?")
        params.append(sales_id)
    if customer_code:
        conditions.append("t_原価見積履歴.客先コード = ?")
        params.append(customer_code)
    if part_no:
        conditions.append("t_原価見積履歴.品番 LIKE ?")
        params.append(f"%{part_no}%")
    if part_name:
        conditions.append("t_原価見積履歴.品名 LIKE ?")
        params.append(f"%{part_name}%")
    if estimate_id:
        conditions.append("t_原価見積履歴.原価見積りID = ?")
        params.append(estimate_id)

    sql = base_sql
    if conditions:
        sql += " WHERE " + " AND ".join(conditions)

    sql += " ORDER BY t_原価見積履歴.原価見積りID"

    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()

        col_names = [c[0] for c in cur.description]
        result_rows = [_estimate_row_for_search_ui(col_names, r) for r in rows]

        conn.close()

        return {"rows": result_rows}

    except Exception as e:
        return {"error": str(e)}


def _normalize_estimate_row_keys(d):
    """カーソル列名のテーブル接頭辞・Access の [] を除き、表示用キーに寄せる"""
    if not d:
        return d
    out = {}
    for k, v in d.items():
        nk = k
        if isinstance(k, str):
            if "." in k:
                nk = k.rsplit(".", 1)[-1]
            if nk.startswith("[") and nk.endswith("]"):
                nk = nk[1:-1].strip()
        out[nk] = v
    return out


_ESTIMATE_SEARCH_UI_KEYS = [
    "原価見積りID",
    "管理NO",
    "営業担当",
    "客先名",
    "品番",
    "品名",
    "備考",
]


def _estimate_row_for_search_ui(col_names, r):
    """検索画面・登録直後表示用の 7 キー dict（ドライバの列名差を吸収）"""
    if r is None:
        return None
    raw = dict(zip(col_names, r))
    norm = _normalize_estimate_row_keys(raw)
    if all(k in norm for k in _ESTIMATE_SEARCH_UI_KEYS):
        return {k: norm[k] for k in _ESTIMATE_SEARCH_UI_KEYS}
    if len(r) >= len(_ESTIMATE_SEARCH_UI_KEYS):
        return {k: r[i] for i, k in enumerate(_ESTIMATE_SEARCH_UI_KEYS)}
    return norm


def _json_safe_estimate_row(d):
    """pyodbc 行 dict を jsonify 可能な型にそろえる"""
    out = {}
    for k, v in d.items():
        if v is None:
            out[k] = None
        elif isinstance(v, (str, int, float, bool)):
            out[k] = v
        else:
            out[k] = str(v)
    return out


def _bikou_newlines_to_crlf(s):
    """備考の改行を Access / VB の Chr(13) & Chr(10) と同じ CRLF にそろえる"""
    if not s:
        return s
    if not isinstance(s, str):
        s = str(s)
    normalized = s.replace("\r\n", "\n").replace("\r", "\n")
    return normalized.replace("\n", "\r\n")


def _fetch_estimate_row_for_search(cur, estimate_id):
    """search_conditions と同じ列構成で 1 行取得"""
    sql = (
        "SELECT "
        "t_原価見積履歴.原価見積りID, "
        "t_原価見積履歴.管理NO, "
        "t_営業マスタ.営業担当, "
        "t_客先マスタ.客先名, "
        "t_原価見積履歴.品番, "
        "t_原価見積履歴.品名, "
        "t_原価見積履歴.備考 "
        "FROM ((t_原価見積履歴 LEFT JOIN t_営業マスタ ON t_原価見積履歴.営業ID = t_営業マスタ.コード) "
        "LEFT JOIN t_客先マスタ ON t_原価見積履歴.客先コード = t_客先マスタ.コード) "
        "WHERE t_原価見積履歴.原価見積りID = ?"
    )
    cur.execute(sql, (estimate_id,))
    r = cur.fetchone()
    if not r:
        return None
    col_names = [c[0] for c in cur.description]
    return _estimate_row_for_search_ui(col_names, r)


def api_register_estimate(payload=None):
    """新規モーダルから t_原価見積履歴 へ 1 件追加し、追加行を返す"""
    data = payload or {}
    sales_id = (data.get('sales_id') or '').strip()
    kanri_no = (data.get('kanri_no') or '').strip()
    customer_code = (data.get('customer_code') or '').strip()
    part_no = (data.get('part_no') or '').strip()
    part_name_raw = data.get('part_name')
    bikou_raw = data.get('bikou')
    part_name = (part_name_raw or '').strip() if part_name_raw is not None else ''
    bikou = (bikou_raw or '').strip() if bikou_raw is not None else ''

    if not sales_id or not kanri_no or not customer_code or not part_no:
        return {"error": "必須項目が不足しています"}

    part_name_val = part_name if part_name else None
    bikou_val = _bikou_newlines_to_crlf(bikou) if bikou else None

    insert_sql = (
        "INSERT INTO t_原価見積履歴 (営業ID, 管理NO, 客先コード, 品番, 品名, 備考) "
        "VALUES (?, ?, ?, ?, ?, ?)"
    )
    params = (sales_id, kanri_no, customer_code, part_no, part_name_val, bikou_val)

    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(insert_sql, params)
        conn.commit()

        cur.execute("SELECT MAX(原価見積りID) FROM t_原価見積履歴")
        max_row = cur.fetchone()
        new_id = max_row[0] if max_row else None
        if new_id is None:
            conn.close()
            return {"error": "登録後の原価見積りIDを取得できませんでした"}

        row = _fetch_estimate_row_for_search(cur, new_id)
        conn.close()

        if not row:
            return {"error": "登録した行を取得できませんでした"}

        return {"row": _json_safe_estimate_row(row)}
    except Exception as e:
        return {"error": str(e)}


def api_update_estimate_history(payload=None):
    """検索画面編集モーダル: t_原価見積履歴 を原価見積りID で更新"""
    data = payload or {}
    estimate_id = (data.get("estimate_id") or "").strip()
    sales_id = (data.get("sales_id") or "").strip()
    kanri_no = (data.get("kanri_no") or "").strip()
    customer_code = (data.get("customer_code") or "").strip()
    part_no = (data.get("part_no") or "").strip()
    part_name_raw = data.get("part_name")
    bikou_raw = data.get("bikou")
    part_name = (part_name_raw or "").strip() if part_name_raw is not None else ""
    bikou = (bikou_raw or "").strip() if bikou_raw is not None else ""

    if not estimate_id:
        return {"error": "原価見積りIDが必要です"}
    if not sales_id or not kanri_no or not customer_code or not part_no:
        return {"error": "必須項目が不足しています"}

    part_name_val = part_name if part_name else None
    bikou_val = _bikou_newlines_to_crlf(bikou) if bikou else None

    update_sql = (
        "UPDATE t_原価見積履歴 SET "
        "営業ID = ?, 管理NO = ?, 客先コード = ?, 品番 = ?, 品名 = ?, 備考 = ? "
        "WHERE 原価見積りID = ?"
    )
    params = (sales_id, kanri_no, customer_code, part_no, part_name_val, bikou_val, estimate_id)

    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(update_sql, params)
        if cur.rowcount == 0:
            conn.close()
            return {"error": "該当する原価見積りIDが見つかりません"}
        conn.commit()
        row = _fetch_estimate_row_for_search(cur, estimate_id)
        conn.close()
        if not row:
            return {"error": "更新後の行を取得できませんでした"}
        return {"row": _json_safe_estimate_row(row)}
    except Exception as e:
        return {"error": str(e)}


def api_search_delete_estimate(payload=None):
    """検索画面編集モーダル: 原価見積りID に紐づくロット周り・履歴・情報・初期費用を削除"""
    data = payload or {}
    estimate_id = (data.get("estimate_id") or "").strip()
    if not estimate_id:
        return {"error": "原価見積りIDが必要です"}

    lot_detail_tables = [
        "t_原価見積材料",
        "t_原価見積真鍮",
        "t_原価見積加工管理",
        "t_原価見積送料",
        "t_原価見積計算チャージ",
    ]

    conn = None
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute(
            "SELECT ロットID FROM t_原価見積情報 WHERE 原価見積りID = ?",
            [estimate_id],
        )
        lot_rows = cur.fetchall()
        lot_ids = []
        for r in lot_rows:
            if not r or r[0] is None:
                continue
            lot_ids.append(r[0])

        for lid in lot_ids:
            for table_name in lot_detail_tables:
                cur.execute(f"DELETE FROM {table_name} WHERE ロットID = ?", [lid])

        cur.execute("DELETE FROM t_原価見積情報 WHERE 原価見積りID = ?", [estimate_id])
        cur.execute("DELETE FROM t_原価見積履歴 WHERE 原価見積りID = ?", [estimate_id])
        cur.execute("DELETE FROM t_原価見積初期費用 WHERE 原価見積りID = ?", [estimate_id])

        conn.commit()
        return {"ok": True}
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        return {"error": str(e)}
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def api_search(payload=None):
    """querytest.html からの検索API"""
    q = ((payload or {}).get('q') or '').strip()

    try:
        conn = get_connection()
        cur = conn.cursor()

        if q:
            # 品名を部分一致で検索（必要に応じて列名を調整してください）
            sql = "SELECT * FROM t_原価見積履歴 WHERE 品名 LIKE ?"
            params = ['%' + q + '%']
        else:
            # 未入力なら全件取得
            sql = "SELECT * FROM t_原価見積履歴"
            params = []

        cur.execute(sql, params)
        rows = cur.fetchall()

        # カラム名を取得し、dict のリストに変換（querytest.html がそのままテーブル化）
        col_names = [c[0] for c in cur.description]
        result_rows = [dict(zip(col_names, r)) for r in rows]

        conn.close()

        return {"rows": result_rows}

    except Exception as e:
        # エラー内容をそのまま返す（開発中）
        return {"error": str(e)}
